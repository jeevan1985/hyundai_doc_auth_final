#!/usr/bin/env python
"""
Optional disk-backed key index using SQLite for scalable, reusable lookups.

This module provides an opt-in, production-grade SQLite-backed index for the
key table used by the key-driven pipeline. It eliminates RAM spikes by avoiding
loading the entire table as an in-memory dict, while supporting O(1)-ish lookups
via indexed columns.

Design highlights
- Backward compatible: only active when key_input.disk_backed_index.enabled=true.
- Flexible ingestion: CSV (stream), NDJSON/JSON, Excel (openpyxl read_only → temp CSV → stream).
- Configurable schema: store the filename key, optional normalized_key for case-insensitive
  lookups, a minimal set of required columns, and optionally the entire source row as row_json.
- Robust rebuild policy: detect source/schema changes and rebuild as needed.
- Casting pipeline: column_types + casting rules normalize values on ingestion.
- Reuse across runs: metadata table records source signature, schema, and build version.

Extensibility notes (LMDB or alternative KV stores)
- The helpers and function boundaries are designed so a future 'backend: lmdb' can implement
  the same public APIs (build, is_up_to_date, lookup) with minimal changes to the orchestrator.
- The ingest/casting/required-columns logic is backend-agnostic and can be shared.

Public API
- build_sqlite_index(config, required_columns) -> None
- is_index_up_to_date(config, required_columns) -> (bool, reason)
- lookup_rows_by_keys(config, keys, required_columns) -> { filename -> {col: value, ...} }
- Helpers to derive required columns/declared types from the YAML config.

All functions are heavily commented to document design trade-offs and failure modes.
"""
from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

LOGGER = logging.getLogger(__name__)
if not LOGGER.handlers:
    LOGGER.addHandler(logging.NullHandler())
LOGGER.propagate = True

INDEX_VERSION = "1.0"

# ==============================================================================
# Small helpers (paths, config access)
# ==============================================================================

def _project_root() -> Path:
    """Return the repository root directory path.

    This resolves the project root relative to this file location.

    Returns:
        Path: Absolute path to the repository root.
    """
    # external/key_input/sqlite_key_index.py -> key_input -> external -> repo root
    return Path(__file__).resolve().parents[2]


def _resolve_db_path(kcfg: Dict[str, Any]) -> Path:
    """Resolve the absolute SQLite database path from configuration.

    Default location for the key index file is now
    'instance/query_key_index/key_table_index.sqlite'.
    An explicit key_input.disk_backed_index.sqlite.db_path still takes precedence.
    If a global search_task.query_embed_index.output_path_query_embed_index is set
    and db_path is not explicitly provided, this function will honor it.
    
    Args:
        kcfg (Dict[str, Any]): Top-level configuration dictionary.

    Returns:
        Path: Absolute path to the SQLite database file; parent directories ensured.
   
    """
    dcfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
    scfg = (dcfg.get("sqlite") or {})
    db_path = scfg.get("db_path")
    if not db_path:
        # Try global override from search_task.query_embed_index
        try:
            qei = (kcfg.get("search_task", {}).get("query_embed_index", {}) or {})
            base_override = qei.get("output_path_query_embed_index")
        except Exception:
            base_override = None
        if base_override:
            db_path = str(Path(base_override) / "key_table_index.sqlite")
        else:
            db_path = "instance/query_key_index/key_table_index.sqlite"
    p = Path(db_path)
    if not p.is_absolute():
        p = _project_root() / p
    p.parent.mkdir(parents=True, exist_ok=True)
    return p.resolve()


def _detect_input_format(input_path: Path, explicit: Optional[str]) -> str:
    """Detect the input table format from extension or explicit override.

    Args:
        input_path (Path): Path to the input table file.
        explicit (Optional[str]): Explicit format ('csv'|'excel'|'json'|'auto').

    Returns:
        str: Detected format string ('csv'|'excel'|'json').
    """
    fmt = (explicit or "auto").lower()
    if fmt == "auto":
        suf = input_path.suffix.lower()
        if suf in (".csv", ".tsv"):
            return "csv"
        if suf in (".xlsx", ".xls"):
            return "excel"
        if suf in (".json", ".ndjson", ".jsonl"):
            return "json"
        return "csv"
    return fmt


# ==============================================================================
# Derivation helpers exposed publicly
# ==============================================================================

def derive_required_columns(kcfg: Dict[str, Any]) -> List[str]:
    """Derive the minimal required set of columns to persist in the index.

    Rules
    - Always include the filename column: key_input.file_name_column
    - Union of:
      - data_source.api.request_mapping.param_map keys (if present)
      - key_input.columns_for_results (if present)
    - If store_all_columns=true, we will additionally store row_json (not a column from the
      source), but that does not change the required column set for schema checks.
    """
    ki = kcfg.get("key_input", {})
    file_name_column = str(ki.get("file_name_column", "파일명"))
    cols: List[str] = [file_name_column]

    # From API param_map
    api_param_map = (
        kcfg.get("data_source", {})
        .get("api", {})
        .get("request_mapping", {})
        .get("param_map", {})
        or {}
    )
    cols.extend(list(api_param_map.keys()))

    # From enrichment
    cols_for_results = ki.get("columns_for_results") or []
    if isinstance(cols_for_results, list):
        cols.extend([str(c) for c in cols_for_results if c])

    # Deduplicate but preserve order (filename first)
    seen = set()
    out: List[str] = []
    for c in cols:
        if c not in seen:
            out.append(c)
            seen.add(c)
    return out


def derive_declared_types(kcfg: Dict[str, Any], required_columns: List[str]) -> Dict[str, str]:
    """Determine declared SQLite types for the required columns.

    Resolves per-column types from key_input.disk_backed_index.sqlite.column_types with fallback
    to key_input.disk_backed_index.sqlite.default_type. Only returns entries for required_columns.
    """
    dcfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
    scfg = (dcfg.get("sqlite") or {})
    default_type = str(scfg.get("default_type", "TEXT")).upper()
    column_types = {str(k): str(v).upper() for k, v in (scfg.get("column_types") or {}).items()}
    declared: Dict[str, str] = {}
    for c in required_columns:
        declared[c] = column_types.get(c, default_type)
    return declared


# ==============================================================================
# Source signature and casting helpers
# ==============================================================================

def compute_source_signature(path: Path, mode: str) -> Dict[str, Any]:
    """Compute a signature for the input source to detect changes between runs.

    Returns a dict including mtime (UTC epoch seconds), size (bytes), and optionally a checksum
    hash when mode == 'mtime_hash'.
    """
    st = path.stat()
    sig: Dict[str, Any] = {
        "source_path": str(path),
        "source_mtime": st.st_mtime,
        "source_size": st.st_size,
    }
    if mode == "mtime_hash":
        # Stream the file to compute SHA256; robust for large inputs.
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        sig["checksum"] = h.hexdigest()
    return sig


def normalize_key(value: str, case_insensitive: bool) -> str:
    """Normalize a filename key for case-insensitive matching when enabled.

    Args:
        value (str): Original key value from source row.
        case_insensitive (bool): Whether to lowercase keys for matching.

    Returns:
        str: Normalized key string.
    """
    return value.lower() if (case_insensitive and isinstance(value, str)) else value


def _parse_datetime(value: Any, input_format: Optional[str]) -> Optional[datetime]:
    """Parse a date/time value from string to datetime if possible.

    Supports ISO-8601 with optional trailing 'Z', or a custom input_format.

    Args:
        value (Any): Input value to parse.
        input_format (Optional[str]): Optional datetime.strptime format string.

    Returns:
        Optional[datetime]: Parsed datetime or None if parsing fails.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        if input_format:
            return datetime.strptime(s, input_format)
        # Try ISO variants; handle trailing Z
        s2 = s[:-1] if s.endswith("Z") else s
        return datetime.fromisoformat(s2)
    except Exception:
        return None


def apply_ingest_cast(value: Any, column_name: str, sqlite_casting_cfg: Dict[str, Any],
                      global_date_policy: str, on_cast_error: str) -> Any:
    """Normalize a source value for insertion into SQLite based on casting rules.

    Supported casting modes
    - int, float: convert numerics; on error follow on_cast_error
    - date_to_iso: parse input using input_format (optional), then format to ISO-8601 (UTC naive)
    - bool: convert common truthy/falsey strings/numbers to 1/0

    global_date_policy is currently used only as a default target for date-like values.
    """
    rule = (sqlite_casting_cfg or {}).get(column_name)
    if not rule:
        return value

    mode = str(rule.get("mode", "")).lower()
    try:
        if mode == "int":
            if value is None or str(value).strip() == "":
                return None
            return int(str(value).strip())
        if mode == "float":
            if value is None or str(value).strip() == "":
                return None
            return float(str(value).replace(",", "").strip())
        if mode == "bool":
            if value is None:
                return None
            s = str(value).strip().lower()
            if s in ("1", "true", "t", "yes", "y"):
                return 1
            if s in ("0", "false", "f", "no", "n"):
                return 0
            # Fallback: non-empty -> true
            return 1 if s else 0
        if mode == "date_to_iso":
            dt = _parse_datetime(value, rule.get("input_format"))
            if not dt:
                return None
            # Default to ISO text; other global policies can be added later
            return dt.replace(tzinfo=None).isoformat(timespec="seconds")
    except Exception:
        # fall through to error policy
        pass

    if on_cast_error == "store_null":
        return None
    if on_cast_error == "raise":
        raise
    # Default: store_text (stringify original)
    return None if value is None else str(value)


# ==============================================================================
# Index lifecycle
# ==============================================================================

def is_index_up_to_date(kcfg: Dict[str, Any], required_columns: List[str]) -> Tuple[bool, str]:
    """Check if the SQLite index exists and matches the current source/schema.

    Returns (True, "ok") when up-to-date, else (False, reason).
    """
    db_path = _resolve_db_path(kcfg)
    if not db_path.exists():
        return False, "db_missing"

    dcfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
    scfg = (dcfg.get("sqlite") or {})
    case_insensitive = bool(dcfg.get("case_insensitive_keys", True))

    # Resolve rebuild policy
    rpol = dcfg.get("rebuild_policy", {}) or {}
    src_mode = str(rpol.get("on_source_change", "mtime_hash")).lower()
    schema_check = bool(rpol.get("on_schema_change", True))

    # Resolve source signature when applicable
    ki = kcfg.get("key_input", {})
    input_table_path = Path(ki.get("input_table_path"))
    src_sig = None
    if src_mode in ("mtime_only", "mtime_hash"):
        src_sig = compute_source_signature(input_table_path, src_mode)

    # Open and read metadata
    try:
        con = sqlite3.connect(str(db_path))
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        # Ensure metadata table exists
        cur.execute("""
            SELECT name FROM sqlite_master WHERE type='table' AND name=?
        """, (scfg.get("metadata_table", "key_index_meta"),))
        row = cur.fetchone()
        if not row:
            return False, "meta_missing"
        meta_table = scfg.get("metadata_table", "key_index_meta")
        cur.execute(f"SELECT * FROM {meta_table} LIMIT 1")
        meta = cur.fetchone()
        if not meta:
            return False, "meta_empty"

        # Compare source signature and schema settings
        def _safe(mkey: str, default: Any = None) -> Any:
            """Safely extract a field from the meta row with a default.

            Args:
                mkey (str): Metadata key.
                default (Any): Fallback value when the key is missing.

            Returns:
                Any: Retrieved value or the default.
            """
            try:
                return meta[mkey]
            except Exception:
                return default

        # Source checks (honor on_source_change policy)
        if src_mode in ("mtime_only", "mtime_hash") and src_sig is not None:
            if abs(float(_safe("source_mtime", 0.0)) - float(src_sig.get("source_mtime", 0.0))) > 1e-6:
                return False, "source_mtime_changed"
            if int(_safe("source_size", 0)) != int(src_sig.get("source_size", 0)):
                return False, "source_size_changed"
            m_checksum = _safe("checksum")
            if src_mode == "mtime_hash" and src_sig.get("checksum") and m_checksum and m_checksum != src_sig.get("checksum"):
                return False, "source_checksum_changed"

        # Schema checks (honor on_schema_change policy)
        if schema_check:
            m_file_col = _safe("file_name_column")
            if m_file_col and m_file_col != ki.get("file_name_column", "파일명"):
                return False, "file_name_column_changed"

            m_case_ins = bool(int(_safe("case_insensitive_keys", 1)))
            if m_case_ins != case_insensitive:
                return False, "case_insensitive_keys_changed"

            m_required = json.loads(_safe("required_columns_json", "[]") or "[]")
            if sorted([str(c) for c in required_columns]) != sorted([str(c) for c in m_required]):
                return False, "required_columns_changed"

            # declared types
            declared = derive_declared_types(kcfg, required_columns)
            m_declared = json.loads(_safe("declared_types_json", "{}") or "{}")
            if {k: v.upper() for k, v in declared.items()} != {k: str(v).upper() for k, v in m_declared.items()}:
                return False, "declared_types_changed"

        # Version mismatch may require rebuild in future; for now accept older versions
        return True, "ok"
    finally:
        try:
            con.close()
        except Exception:
            pass


def build_sqlite_index(kcfg: Dict[str, Any], required_columns: List[str]) -> None:
    """Build or rebuild the SQLite index according to configuration.

    Rebuild triggers are determined by the orchestrator using is_index_up_to_date() and
    rebuild_policy.force_rebuild; this function always rebuilds the DB file fresh.

    Steps
    - Resolve schema: columns + types + key + normalized_key
    - Stream input rows and insert in chunks (transactions)
    - Persist metadata (source signature, schema, etc.)
    - Apply PRAGMA settings (WAL, synchronous)
    """
    dcfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
    scfg = (dcfg.get("sqlite") or {})

    db_path = _resolve_db_path(kcfg)
    table_name = scfg.get("table_name", "key_index")
    meta_table = scfg.get("metadata_table", "key_index_meta")
    journal_mode = scfg.get("journal_mode", "WAL")
    synchronous = scfg.get("synchronous", "NORMAL")
    strict_mode = bool(scfg.get("strict_mode", False))
    default_type = str(scfg.get("default_type", "TEXT")).upper()
    column_types = derive_declared_types(kcfg, required_columns)

    index_chunk_size = int(dcfg.get("index_chunk_size", 10000))
    store_all_columns = bool(dcfg.get("store_all_columns", False))
    case_insensitive = bool(dcfg.get("case_insensitive_keys", True))

    casting_cfg = (scfg.get("casting") or {})
    on_cast_error = str(scfg.get("on_cast_error", "store_text")).lower()
    date_policy = str(scfg.get("date_policy", "iso_text")).lower()

    ki = kcfg.get("key_input", {})
    file_name_column = str(ki.get("file_name_column", "파일명"))
    input_table_path = Path(ki.get("input_table_path"))
    input_format = _detect_input_format(input_table_path, ki.get("format"))

    # Announce
    LOGGER.info("[disk-index] build start: backend=sqlite db=%s table=%s chunk=%s journal=%s sync=%s",
                db_path, table_name, index_chunk_size, journal_mode, synchronous)
    LOGGER.info("[disk-index] required columns: %s", ", ".join(required_columns))
    LOGGER.info("[disk-index] declared types: %s", ", ".join(f"{k}:{v}" for k, v in column_types.items()))

    # Prepare schema DDL
    col_defs = ["key TEXT NOT NULL UNIQUE"]
    if case_insensitive:
        col_defs.append("normalized_key TEXT")
    for c in required_columns:
        # Avoid duplicating the key column (we still store it explicitly to preserve the raw source value)
        ctype = column_types.get(c, default_type)
        # Sanitize column name for SQLite (simple quote); assume names are valid identifiers otherwise.
        col_defs.append(f"\"{c}\" {ctype}")
    if store_all_columns:
        col_defs.append("row_json TEXT")
    col_defs_str = ", ".join(col_defs)

    con = sqlite3.connect(str(db_path))
    try:
        cur = con.cursor()
        # PRAGMAs
        try:
            cur.execute(f"PRAGMA journal_mode={journal_mode}")
        except Exception:
            pass
        try:
            cur.execute(f"PRAGMA synchronous={synchronous}")
        except Exception:
            pass

        # Drop existing
        cur.execute(f"DROP TABLE IF EXISTS {table_name}")
        cur.execute(f"DROP TABLE IF EXISTS {meta_table}")

        # Create tables
        # TODO: Validate table/column identifiers from configuration to mitigate
        #       SQL injection risks in dynamic identifier usage.
        ddl = f"CREATE TABLE {table_name} ({col_defs_str})"
        if strict_mode:
            # Try STRICT mode (SQLite 3.37+). If unsupported, fallback silently.
            try:
                cur.execute(ddl + " STRICT")
            except Exception:
                cur.execute(ddl)
        else:
            cur.execute(ddl)
        if case_insensitive:
            cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{table_name}_nkey ON {table_name}(normalized_key)")

        cur.execute(
            f"""
            CREATE TABLE {meta_table} (
                source_path TEXT,
                source_mtime REAL,
                source_size INTEGER,
                checksum TEXT,
                file_name_column TEXT,
                case_insensitive_keys INTEGER,
                required_columns_json TEXT,
                declared_types_json TEXT,
                build_time_utc TEXT,
                version TEXT
            )
            """
        )
        con.commit()

        # Stream source rows and insert in chunks
        def iter_rows() -> Iterator[Dict[str, Any]]:
            """Yield source rows from CSV, Excel, or JSON in a streaming fashion."""
            if input_format == "csv":
                with input_table_path.open("r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        yield r
            elif input_format == "excel":
                try:
                    from openpyxl import load_workbook  # type: ignore
                except ImportError as e:
                    raise RuntimeError(
                        "Excel ingestion requires 'openpyxl'. Install it or convert to CSV."
                    ) from e
                # Convert to temp CSV for streaming
                with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv", encoding="utf-8") as tmp_f:
                    tmp_csv = Path(tmp_f.name)
                wb = load_workbook(filename=str(input_table_path), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header = ["" if h is None else str(h) for h in next(rows_iter)]
                except StopIteration:
                    wb.close()
                    return
                with tmp_csv.open("w", encoding="utf-8", newline="") as fcsv:
                    w = csv.writer(fcsv)
                    w.writerow(header)
                    for r in rows_iter:
                        w.writerow(["" if c is None else c for c in r])
                wb.close()
                with tmp_csv.open("r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        yield r
                try:
                    tmp_csv.unlink()
                except Exception:
                    pass
            else:  # json
                with input_table_path.open("r", encoding="utf-8") as f:
                    suf = input_table_path.suffix.lower()
                    if suf in (".ndjson", ".jsonl"):
                        for line in f:
                            line = line.strip()
                            if not line:
                                continue
                            try:
                                obj = json.loads(line)
                            except json.JSONDecodeError:
                                continue
                            if isinstance(obj, dict):
                                yield obj
                    else:
                        data = json.load(f)
                        if isinstance(data, dict):
                            # allow object with nested arrays
                            rows = []
                            for _k, v in data.items():
                                if isinstance(v, list):
                                    rows.extend(v)
                            data = rows
                        if not isinstance(data, list):
                            raise RuntimeError("JSON must be an array of objects for ingestion.")
                        for obj in data:
                            if isinstance(obj, dict):
                                yield obj

        # Validate required columns exist in header/first row (best effort for JSON ND)
        first_row: Optional[Dict[str, Any]] = None
        for r in iter_rows():
            first_row = r
            break
        if first_row is None:
            LOGGER.warning("[disk-index] source appears empty; creating empty index")
            rows_iter: Iterator[Dict[str, Any]] = iter(())
        else:
            missing = [c for c in required_columns if c not in first_row]
            if missing:
                raise RuntimeError(f"Missing required columns in source: {missing}")
            # Recreate generator including the first row
            def gen_with_first(fr: Dict[str, Any]) -> Iterator[Dict[str, Any]]:
                """Yield a first row followed by the main iterator stream."""
                yield fr
                for r in iter_rows():
                    yield r
            rows_iter = gen_with_first(first_row)

        # Prepare insert statement
        cols_for_insert = ["key"]
        if case_insensitive:
            cols_for_insert.append("normalized_key")
        cols_for_insert.extend([f"\"{c}\"" for c in required_columns])
        if store_all_columns:
            cols_for_insert.append("row_json")
        placeholders = ",".join(["?"] * len(cols_for_insert))
        insert_sql = f"INSERT OR REPLACE INTO {table_name} ({','.join(cols_for_insert)}) VALUES ({placeholders})"

        start_ts = datetime.now(tz=timezone.utc)
        batch: List[Tuple[Any, ...]] = []
        total_rows = 0
        for row in rows_iter:
            key_val = row.get(file_name_column)
            if key_val is None:
                # Skip rows lacking a key
                continue
            key_val = str(key_val)
            nkey = normalize_key(key_val, case_insensitive)
            values: List[Any] = [key_val]
            if case_insensitive:
                values.append(nkey)
            # Per-column casting
            for c in required_columns:
                v = row.get(c)
                v2 = apply_ingest_cast(v, c, casting_cfg, date_policy, on_cast_error)
                values.append(v2)
            if store_all_columns:
                try:
                    values.append(json.dumps(row, ensure_ascii=False))
                except Exception:
                    values.append(None)
            batch.append(tuple(values))
            if len(batch) >= index_chunk_size:
                cur.executemany(insert_sql, batch)
                con.commit()
                total_rows += len(batch)
                batch = []
        if batch:
            cur.executemany(insert_sql, batch)
            con.commit()
            total_rows += len(batch)

        # Metadata
        src_mode = str(dcfg.get("rebuild_policy", {}).get("on_source_change", "mtime_hash")).lower()
        sig = compute_source_signature(input_table_path, src_mode if src_mode in ("mtime_only", "mtime_hash") else "mtime_hash")
        cur.execute(
            f"INSERT INTO {meta_table} (source_path, source_mtime, source_size, checksum, file_name_column, "
            f"case_insensitive_keys, required_columns_json, declared_types_json, build_time_utc, version) "
            f"VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                sig.get("source_path"),
                sig.get("source_mtime"),
                sig.get("source_size"),
                sig.get("checksum"),
                file_name_column,
                1 if case_insensitive else 0,
                json.dumps(required_columns, ensure_ascii=False),
                json.dumps(column_types, ensure_ascii=False),
                datetime.now(tz=timezone.utc).isoformat(timespec="seconds"),
                INDEX_VERSION,
            ),
        )
        con.commit()

        dur = (datetime.now(tz=timezone.utc) - start_ts).total_seconds() or 1.0
        rps = int(total_rows / dur)
        LOGGER.info("[disk-index] build complete: rows=%d, time=%.1fs, rows/sec=%d", total_rows, dur, rps)
    finally:
        try:
            con.close()
        except Exception:
            pass


def lookup_rows_by_keys(kcfg: Dict[str, Any], keys: List[str], required_columns: List[str]) -> Dict[str, Dict[str, Any]]:
    """Lookup a batch of keys and return a mapping filename -> row dict.

    - case_insensitive_keys=true uses normalized_key for matching
    - Keys without a match are omitted from the result
    - Returned row dicts contain only the required columns plus the filename column (same as input names)
    """
    if not keys:
        return {}
    dcfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
    scfg = (dcfg.get("sqlite") or {})
    case_insensitive = bool(dcfg.get("case_insensitive_keys", True))
    file_name_column = str(kcfg.get("key_input", {}).get("file_name_column", "파일명"))

    db_path = _resolve_db_path(kcfg)
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row
    try:
        cur = con.cursor()
        table_name = scfg.get("table_name", "key_index")

        # We need to map original keys back. For case-insensitive, we compare lower(key).
        # Prepare query columns
        select_cols = [f'"{file_name_column}"'] + [f'"{c}"' for c in required_columns if c != file_name_column]
        select_cols_sql = ",".join(select_cols)

        # SQLite has a default limit of 999 vars per statement - chunk accordingly
        max_vars = 900
        result: Dict[str, Dict[str, Any]] = {}
        for i in range(0, len(keys), max_vars):
            chunk = keys[i : i + max_vars]
            if case_insensitive:
                nkeys = [normalize_key(k, True) for k in chunk]
                placeholders = ",".join(["?"] * len(nkeys))
                sql = f"SELECT key, {select_cols_sql} FROM {table_name} WHERE normalized_key IN ({placeholders})"
                cur.execute(sql, nkeys)
            else:
                placeholders = ",".join(["?"] * len(chunk))
                sql = f"SELECT key, {select_cols_sql} FROM {table_name} WHERE key IN ({placeholders})"
                cur.execute(sql, chunk)

            for row in cur.fetchall():
                key_val = row["key"]
                row_dict = {file_name_column: row[file_name_column]}
                for c in required_columns:
                    if c == file_name_column:
                        continue
                    try:
                        row_dict[c] = row[c]
                    except Exception:
                        row_dict[c] = None
                result[str(key_val)] = row_dict
        return result
    finally:
        try:
            con.close()
        except Exception:
            pass
