#!/usr/bin/env python
"""
Key-driven TIF pipeline orchestration (optional helper module)

Purpose
- Provide a helper to obtain TIF images in batches based on a key table (CSV/Excel/JSON) that
contains filenames (e.g., column "파일명").
- Support two data sources:
- local: resolve filenames to files under configured search roots
- database: fetch file paths or blobs from a PostgreSQL/EDB database
- For each batch, run the existing TIF batch workflow exactly as before (no changes to core_engine).

Notes
- This module intentionally does NOT introduce new pipeline knobs. It uses the existing
image_similarity_config.yaml as-is. Any pipeline behavior remains under that configuration
or CLI arguments in the calling script.
- If doc_input_start != 'key', this module is a no-op.

Author: JK Image Similarity System contributors
"""
from __future__ import annotations

import csv
import json
import logging
import os
import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple, Union, TYPE_CHECKING

# Disk-backed index (optional)
try:
    from .sqlite_key_index import (
        derive_required_columns,
        derive_declared_types,
        is_index_up_to_date,
        build_sqlite_index,
        lookup_rows_by_keys,
    )
except Exception:
    # Keep import-time failures non-fatal for backward compatibility
    derive_required_columns = None  # type: ignore
    derive_declared_types = None  # type: ignore
    is_index_up_to_date = None  # type: ignore
    build_sqlite_index = None  # type: ignore
    lookup_rows_by_keys = None  # type: ignore

import yaml
import base64
import concurrent.futures
import time
import requests
from datetime import datetime, timezone
import threading

# Optional failed-request JSONL logger
try:
    from core_engine.image_similarity_system.log_utils import log_failed_key_request  # type: ignore
except Exception:  # pragma: no cover - resilience if module relocation
    def log_failed_key_request(*args, **kwargs):  # type: ignore
        return False

# Optional log maintenance utility (defensive import)
try:
    from core_engine.image_similarity_system.utils import maintain_log_files  # type: ignore
except Exception:  # pragma: no cover - resilience if module relocation
    def maintain_log_files(*_args, **_kwargs):  # type: ignore
        return 0

# Optional imports (available in project)
try:
    from core_engine.image_similarity_system.config_loader import load_and_merge_configs
except ImportError:
    load_and_merge_configs = None # Fallback to raw YAML if unavailable

try:
    from core_engine.image_similarity_system.workflow import execute_tif_batch_search_workflow, build_index_from_tif_folder_workflow
except ImportError:
    execute_tif_batch_search_workflow = None # type: ignore
    build_index_from_tif_folder_workflow = None # type: ignore

# ==============================================================================
# Utilities
# ==============================================================================
def _get_logger(name: str = __name__) -> logging.Logger:
    """Create a namespaced logger without configuring handlers.

    The returned logger uses a NullHandler to avoid "No handler found" warnings
    when the host application hasn't configured logging yet. Propagation is
    enabled so upstream handlers can capture messages.

    Args:
        name (str): Logger name. Defaults to the current module's ``__name__``.

    Returns:
        logging.Logger: Logger instance with a NullHandler and propagation enabled.
    """
    logger = logging.getLogger(name)
    if not logger.handlers:
        logger.addHandler(logging.NullHandler())
    logger.propagate = True
    return logger

def _ensure_iterable_chunks(items: Iterable[str], chunk_size: int) -> Iterator[List[str]]:
    """Yield fixed-size chunks from an input iterable of strings.

    Args:
        items (Iterable[str]): Source iterable of string items.
        chunk_size (int): Maximum size for each emitted chunk (must be >= 1).

    Yields:
        Iterator[List[str]]: Lists of up to ``chunk_size`` items preserving order.

    Raises:
        ValueError: If ``chunk_size`` is less than 1.
    """
    if chunk_size < 1:
        raise ValueError("chunk_size must be >= 1")
    batch: List[str] = []
    for item in items:
        batch.append(item)
        if len(batch) >= chunk_size:
            yield batch
            batch = []
    if batch:
        yield batch

# ==============================================================================
# Key Input Loader (streaming filenames from CSV / Excel / JSON)
# ==============================================================================
@dataclass
class KeyInputConfig:
    """Configuration for key table ingestion and batching.

    Attributes:
        input_table_path (Path): Path to CSV/Excel/JSON key table.
        file_name_column (str): Column name containing filenames.
        format (str): 'csv'|'excel'|'json'|'auto'.
        json_array_field (Optional[str]): If JSON wraps an array under a field.
        json_records_is_lines (bool): NDJSON/JSONL streaming mode.
        batch_size (int): Batch size for processing.
        deduplicate (bool): Drop duplicate filenames within a run.
        strip_whitespace (bool): Trim whitespace from filenames.
        key_in_xlsx_to_csv (bool): Stream Excel via temp CSV when possible.
        excel_sheet_name (Optional[Union[str, int]]): Sheet name or index.
        columns_for_results (Optional[List[str]]): Extra columns to enrich results.
    """
    input_table_path: Path
    file_name_column: str = "파일명"
    format: str = "auto" # csv | excel | json | auto
    json_array_field: Optional[str] = None
    json_records_is_lines: bool = False
    batch_size: int = 200
    deduplicate: bool = True
    strip_whitespace: bool = True
    # New options for Excel→CSV streaming
    key_in_xlsx_to_csv: bool = True
    excel_sheet_name: Optional[Union[str, int]] = None
        # New: columns to carry into results
    columns_for_results: Optional[List[str]] = None

class KeyInputLoader:
    """Streams filenames from a key input table.

    Supports CSV (streaming), Excel (with pandas/openpyxl; may load in memory),
    and JSON (array or NDJSON).

    Large Excel files may not be memory-efficient; users are encouraged to use CSV
    for scalable streaming. Raises descriptive errors when optional dependencies are
    missing.
    """
    def __init__(self, config: KeyInputConfig, logger: Optional[logging.Logger] = None) -> None:
        self.config = config
        self.logger = logger or _get_logger(__name__ + ".KeyInputLoader")

    def iter_filenames(self) -> Iterator[str]:
        """Iterate filenames from the configured key table in a streaming manner.

        Auto-detects format when ``format='auto'`` based on file extension and
        dispatches to the appropriate reader. Supported formats are CSV, Excel,
        and JSON (array or NDJSON/JSONL).

        Yields:
            Iterator[str]: Canonicalized filename strings from the key table.

        Raises:
            ValueError: If the configured format is unsupported.
        """
        fmt = (self.config.format or "auto").lower()
        path = self.config.input_table_path
        if fmt == "auto":
            if path.suffix.lower() in (".csv", ".tsv"):
                fmt = "csv"
            elif path.suffix.lower() in (".xlsx", ".xls"):
                fmt = "excel"
            elif path.suffix.lower() in (".json", ".ndjson", ".jsonl"):
                fmt = "json"
            else:
                fmt = "csv"

        if fmt == "csv":
            yield from self._iter_csv()
        elif fmt == "excel":
            yield from self._iter_excel()
        elif fmt == "json":
            yield from self._iter_json()
        else:
            raise ValueError(f"Unsupported format: {fmt}")

    def _iter_csv_from_path(self, path: Path) -> Iterator[str]:
        """Stream filenames from a CSV file.

        Args:
            path (Path): Path to the CSV file.

        Yields:
            Iterator[str]: Filenames from the configured column, optionally stripped.
        """
        col = self.config.file_name_column
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                raw = row.get(col)
                if raw is None:
                    continue
                val = raw.strip() if self.config.strip_whitespace else raw
                if val:
                    yield val

    def _iter_csv(self) -> Iterator[str]:
        """Stream filenames from the configured CSV path."""
        yield from self._iter_csv_from_path(self.config.input_table_path)

    def _iter_excel(self) -> Iterator[str]:
        """Stream filenames from an Excel file using a memory-safe approach.

        When ``key_in_xlsx_to_csv`` is True, performs a lightweight Excel→CSV
        conversion and streams rows from the temporary CSV. Falls back to using
        pandas for smaller files when streaming is disabled.

        Yields:
            Iterator[str]: Filenames from the configured column.

        Raises:
            RuntimeError: If required optional dependencies for Excel reading are missing.
            IndexError: If a specified sheet index is out of range.
            KeyError: If the configured filename column is not found.
        """
        # If configured, convert to CSV using openpyxl streaming and then stream from CSV
        if self.config.key_in_xlsx_to_csv:
            try:
                from openpyxl import load_workbook # type: ignore
            except ImportError as e: # pragma: no cover
                raise RuntimeError(
                    "Excel→CSV streaming requires 'openpyxl'. Install it or set key_in_xlsx_to_csv=false and convert manually."
                ) from e
            import tempfile

            with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".csv", encoding="utf-8") as tmp_f:
                tmp_csv = Path(tmp_f.name)
            
            self.logger.info("Converting Excel to CSV for streaming: '%s'", tmp_csv)

            wb = load_workbook(filename=str(self.config.input_table_path), read_only=True, data_only=True)
            
            sheet = self.config.excel_sheet_name
            if sheet is None:
                ws = wb.active
            elif isinstance(sheet, int):
                try:
                    ws = wb.worksheets[sheet]
                except Exception as e:
                    wb.close()
                    raise IndexError(f"Excel sheet index out of range: {sheet}") from e
            else:
                ws = wb[sheet]

            # Determine header from first row
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return

            # Write CSV
            with tmp_csv.open("w", encoding="utf-8", newline="") as fcsv:
                w = csv.writer(fcsv)
                w.writerow(header)
                for r in rows_iter:
                    w.writerow(["" if c is None else c for c in r])
            wb.close()

            # Stream from the temp CSV
            yield from self._iter_csv_from_path(tmp_csv)
            try:
                tmp_csv.unlink()
            except Exception:
                pass
            return

        # Fallback: load via pandas into memory (smaller files only)
        col = self.config.file_name_column
        try:
            import pandas as pd # type: ignore
        except ImportError as e: # pragma: no cover
            raise RuntimeError(
                "Excel support requires pandas and openpyxl. Install them or convert your file to CSV."
            ) from e

        self.logger.warning(
            "Loading Excel into memory; for very large files, consider converting to CSV for streaming."
        )
        df_obj = pd.read_excel(self.config.input_table_path, sheet_name=self.config.excel_sheet_name)
        # When sheet_name=None, pandas returns a dict of DataFrames. Pick the first sheet by default.
        if isinstance(df_obj, dict):
            if not df_obj:
                return
            df = next(iter(df_obj.values()))
        else:
            df = df_obj
        if col not in df.columns:
            raise KeyError(f"Column '{col}' not found in Excel file.")

        for v in df[col].astype(str).tolist():
            val = v.strip() if self.config.strip_whitespace else v
            if val:
                yield val

    def _iter_json(self) -> Iterator[str]:
        """Stream filenames from a JSON array or NDJSON/JSONL file.

        Yields:
            Iterator[str]: Filenames from each JSON object where the configured
            column is present and non-empty.

        Raises:
            ValueError: If the JSON structure is not an array or requires a missing field.
        """
        col = self.config.file_name_column
        p = self.config.input_table_path
        if self.config.json_records_is_lines or p.suffix.lower() in (".ndjson", ".jsonl"):
            with p.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    raw = obj.get(col)
                    if raw is None:
                        continue
                    val = raw.strip() if self.config.strip_whitespace else raw
                    if val:
                        yield val
        else:
            with p.open("r", encoding="utf-8") as f:
                data = json.load(f)
                if self.config.json_array_field:
                    data = data.get(self.config.json_array_field, [])

                if isinstance(data, dict):
                    rows = []
                    for _k, v in data.items():
                        if isinstance(v, list):
                            rows.extend(v)
                    data = rows

                if not isinstance(data, list):
                    raise ValueError("JSON must be an array of objects or provide 'json_array_field'.")

                for obj in data:
                    if not isinstance(obj, dict):
                        continue
                    raw = obj.get(col)
                    if raw is None:
                        continue
                    val = raw.strip() if self.config.strip_whitespace else raw
                    if val:
                        yield val

    def load_rows_map(self) -> Dict[str, Dict[str, Any]]:
        """Loads the entire key table into a dictionary mapping filename -> row."""
        fmt = (self.config.format or "auto").lower()
        path = self.config.input_table_path
        if fmt == "auto":
            if path.suffix.lower() in (".csv", ".tsv"):
                fmt = "csv"
            elif path.suffix.lower() in (".xlsx", ".xls"):
                fmt = "excel"
            elif path.suffix.lower() in (".json", ".ndjson", ".jsonl"):
                fmt = "json"
            else:
                fmt = "csv"

        if fmt == "csv":
            return self._load_csv_map()
        elif fmt == "excel":
            return self._load_excel_map()
        elif fmt == "json":
            return self._load_json_map()
        else:
            raise ValueError(f"Unsupported format: {fmt}")

    def _load_csv_map(self) -> Dict[str, Dict[str, Any]]:
        rows_map: Dict[str, Dict[str, Any]] = {}
        col = self.config.file_name_column
        with self.config.input_table_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = row.get(col)
                if key:
                    rows_map[key] = row
        return rows_map

    def _load_excel_map(self) -> Dict[str, Dict[str, Any]]:
        try:
            import pandas as pd
        except ImportError as e:
            raise RuntimeError("Excel support requires pandas. Please install it.") from e

        rows_map: Dict[str, Dict[str, Any]] = {}
        col = self.config.file_name_column
        df_obj = pd.read_excel(self.config.input_table_path, sheet_name=self.config.excel_sheet_name)
        # When sheet_name=None, pandas returns a dict of DataFrames. Pick the first sheet by default.
        if isinstance(df_obj, dict):
            if not df_obj:
                return rows_map
            df = next(iter(df_obj.values()))
        else:
            df = df_obj
        if col not in df.columns:
            raise KeyError(f"Column '{col}' not found in Excel file.")

        for _, row in df.iterrows():
            key = row.get(col)
            if key:
                rows_map[str(key)] = row.to_dict()
        return rows_map

    def _load_json_map(self) -> Dict[str, Dict[str, Any]]:
        rows_map: Dict[str, Dict[str, Any]] = {}
        col = self.config.file_name_column
        p = self.config.input_table_path
        with p.open("r", encoding="utf-8") as f:
            if self.config.json_records_is_lines or p.suffix.lower() in (".ndjson", ".jsonl"):
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                        key = obj.get(col)
                        if key:
                            rows_map[key] = obj
                    except json.JSONDecodeError:
                        continue
            else:
                data = json.load(f)
                if self.config.json_array_field:
                    data = data.get(self.config.json_array_field, [])
                if not isinstance(data, list):
                    raise ValueError("JSON must be an array of objects or provide 'json_array_field'.")
                for obj in data:
                    if isinstance(obj, dict):
                        key = obj.get(col)
                        if key:
                            rows_map[key] = obj
        return rows_map

# ==============================================================================
# Fetchers
# ==============================================================================
@dataclass
class ApiConfig:
    """Configuration for API-based fetching of TIF images.

    Attributes:
        api_endpoint (str): API URL.
        http_method (str): 'GET' or 'POST'.
        headers (Dict[str,str]): Request headers.
        timeout_seconds (int): Request timeout.
        max_retries (int): Retries per request.
        backoff_seconds (int): Delay between retries.
        max_concurrency (int): Parallel request workers.
        param_map (Dict[str,str]): Mapping of source columns to API params.
        api_filename_param (Optional[str]): Explicit API parameter for filename.
        send_mapped_filename (bool): Whether to send mapped filenames.
        param_cast (Dict[str, Any]): Per-parameter casting rules.
        sqlite_declared_types (Dict[str, str]): SQLite type hints.
        disk_index_enabled (bool): Whether disk-backed index is used.
        persist_downloads (bool): Save API downloads persistently.
        transient_download_root (Optional[Path]): Root for transient downloads.
        file_name_column (str): Source filename column name.
        image_payload_type (str): 'base64'|'url'|'binary'.
        image_field (str): JSON field for base64 image.
        url_field (str): JSON field for URL.
        binary_field (str): JSON field for binary payloads.
        file_name_field (Optional[str]): JSON field with returned filename.
        api_output_save_path (Path): Output folder for downloads.
    """
    api_endpoint: str
    http_method: str = "POST"
    headers: Dict[str, str] = field(default_factory=dict)
    timeout_seconds: int = 30
    max_retries: int = 3
    backoff_seconds: int = 2
    max_concurrency: int = 4
    # Map key table columns to API request params
    param_map: Dict[str, str] = field(default_factory=dict)
    api_filename_param: Optional[str] = None # Explicit param for filename
    send_mapped_filename: bool = False  # If true, send mapped filename instead of original in API payload
    # Optional outbound casting overrides per API parameter name
    param_cast: Dict[str, Any] = field(default_factory=dict)
    # Declared SQLite types per source column, used as default hints when disk-backed index is enabled
    sqlite_declared_types: Dict[str, str] = field(default_factory=dict)
    # Whether disk-backed index is enabled (controls default casting behavior)
    disk_index_enabled: bool = False
    # Persistence controls for downloaded images
    persist_downloads: bool = False
    transient_download_root: Optional[Path] = None
    file_name_column: str = "파일명"
    # How to get image data from response
    image_payload_type: str = "base64" # base64 | url | binary
    image_field: str = "image_b64"
    url_field: str = "image_url"
    binary_field: str = "image_bytes"
    file_name_field: Optional[str] = None # If API returns a different filename
    api_output_save_path: Path = Path("./api_fetched_images")

class ApiFetcher:
    """Fetches TIF files from a configurable API endpoint."""
    def __init__(self, cfg: ApiConfig, logger: Optional[logging.Logger] = None, name_map: Optional[NameMappingConfig] = None) -> None:
        self.cfg = cfg
        self.logger = logger or _get_logger(__name__ + ".ApiFetcher")
        self.name_map = name_map or NameMappingConfig()
        self._transient_dir: Optional[Path] = None
        # Feature flag to control failed-request logging; set by orchestrator (default True)
        self.enable_failed_key_request_logging: bool = True
        # Thread-safe storage for final failure details per key (populated in request stage)
        self._fail_lock: threading.Lock = threading.Lock()
        self._last_fail_info: Dict[str, Dict[str, Any]] = {}

    def _log_failure(
        self,
        key_name: str,
        status: Optional[int],
        reason: str,
        corr: Optional[str] = None,
        context: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Emit a JSONL entry describing a failed API filename request.

        Args:
            key_name (str): Requested filename key.
            status (Optional[int]): HTTP status code if known.
            reason (str): Reason or error string.
            corr (Optional[str]): Correlation identifier from headers.
            context (Optional[Dict[str, Any]]): Optional metadata.
        """
        try:
            log_failed_key_request(
                requested_name=key_name,
                api_endpoint=self.cfg.api_endpoint,
                status_code=status,
                reason=reason,
                correlation_id=corr,
                context=context,
                enabled=getattr(self, "enable_failed_key_request_logging", True),
            )
        except Exception:
            # Logging must never break the pipeline
            pass

    def _record_last_failure(
        self, key_name: str, status: Optional[int], reason: str, corr: Optional[str]
    ) -> None:
        with self._fail_lock:
            self._last_fail_info[key_name] = {
                "status_code": status,
                "reason": reason,
                "correlation_id": corr,
            }

    def _pop_last_failure(self, key_name: str) -> Optional[Dict[str, Any]]:
        with self._fail_lock:
            return self._last_fail_info.pop(key_name, None)

    def _cast_outbound(self, api_param: str, src_col: str, value: Any) -> Any:
        """Apply request-time param_cast overrides, else infer types from SQLite when enabled.

        Precedence:
        1) request_mapping.param_cast[api_param] if present
        2) sqlite_declared_types[src_col] when disk-backed index is enabled
        3) passthrough
        """
        # 1) Per-parameter override
        rule = (self.cfg.param_cast or {}).get(api_param)
        if rule:
            t = str(rule.get("type", "")).lower()
            on_err = rule.get("on_error", "keep")
            try:
                if t in ("string", "str"):
                    return None if value is None else str(value)
                if t in ("int", "integer"):
                    return None if value in (None, "") else int(str(value))
                if t in ("float", "number", "real"):
                    return None if value in (None, "") else float(str(value).replace(",", ""))
                if t in ("bool", "boolean"):
                    if value is None:
                        return None
                    s = str(value).strip().lower()
                    if s in ("1", "true", "t", "yes", "y"):
                        return True
                    if s in ("0", "false", "f", "no", "n"):
                        return False
                    return bool(s)
                if t in ("datetime", "timestamp"):
                    inp_fmt = rule.get("input_format")
                    as_mode = str(rule.get("as", "iso")).lower()
                    dt = None
                    if isinstance(value, datetime):
                        dt = value
                    else:
                        s = None if value is None else str(value)
                        if s:
                            try:
                                if inp_fmt:
                                    dt = datetime.strptime(s, inp_fmt)
                                else:
                                    s2 = s[:-1] if s.endswith("Z") else s
                                    dt = datetime.fromisoformat(s2)
                            except Exception:
                                dt = None
                    if not dt:
                        return None if on_err in (None, "null", "none") else value
                    # Normalize to UTC naive
                    if dt.tzinfo is not None:
                        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
                    if as_mode in ("epoch_seconds", "epoch"):
                        return int(dt.replace(tzinfo=timezone.utc).timestamp())
                    if as_mode in ("epoch_millis", "millis"):
                        return int(dt.replace(tzinfo=timezone.utc).timestamp() * 1000)
                    return dt.isoformat(timespec="seconds")
            except Exception:
                return None if on_err in (None, "null", "none") else value

        # 2) Default inference from SQLite declared types when disk index is enabled
        if self.cfg.disk_index_enabled:
            dtype = (self.cfg.sqlite_declared_types or {}).get(src_col, "TEXT").upper()
            try:
                if dtype in ("INTEGER",):
                    return None if value in (None, "") else int(value)
                if dtype in ("REAL", "NUMERIC"):
                    return None if value in (None, "") else float(value)
                if dtype in ("BOOLEAN",):
                    if value is None:
                        return None
                    if isinstance(value, (int, float)):
                        return bool(value)
                    s = str(value).strip().lower()
                    if s in ("1", "true", "t", "yes", "y"):
                        return True
                    if s in ("0", "false", "f", "no", "n"):
                        return False
                    return bool(s)
                if dtype in ("DATETIME", "TIMESTAMP"):
                    if isinstance(value, datetime):
                        dt = value
                    else:
                        dt = None
                        s = None if value is None else str(value)
                        if s:
                            try:
                                s2 = s[:-1] if s.endswith("Z") else s
                                dt = datetime.fromisoformat(s2)
                            except Exception:
                                dt = None
                    if not dt:
                        return value
                    if dt.tzinfo is not None:
                        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
                    return dt.isoformat(timespec="seconds")
            except Exception:
                return value
        # 3) default passthrough
        return value

    def _mapped_core(self, filename: str) -> Optional[str]:
        if not self.name_map.enabled:
            return None
        n = len(filename)
        t = max(0, int(self.name_map.tail_len))
        if n <= t:
            return None
        prefix = filename[: n - t]
        suffix = filename[n - t :]
        return f"{prefix}{self.name_map.insert_token}{suffix}"

    def _prepare_request_payload(self, key_name: str, row: Dict[str, Any]) -> Dict[str, Any]:
        """Build an outbound request payload for a single key table row.

        Applies per-parameter casting rules and honors the filename mapping flag
        for the filename column or explicit ``api_filename_param``.

        Args:
            key_name (str): The filename key from the key table.
            row (Dict[str, Any]): The full row mapping for this key.

        Returns:
            Dict[str, Any]: JSON-serializable payload for the API request.
        """
        payload: Dict[str, Any] = {}
        mapped_name = self._mapped_core(key_name)
        name_to_send = mapped_name if (self.cfg.send_mapped_filename and mapped_name) else key_name

        # Build payload from param_map (apply filename mapping flag if mapping the filename column)
        for row_col, api_param in self.cfg.param_map.items():
            if row_col == self.cfg.file_name_column:
                payload[api_param] = name_to_send
            else:
                raw_val = row.get(row_col)
                payload[api_param] = self._cast_outbound(api_param, row_col, raw_val)
        
        # Explicitly set filename parameter, honoring send_mapped_filename
        if self.cfg.api_filename_param:
            payload[self.cfg.api_filename_param] = name_to_send
        # Legacy behavior: if filename column is not in param_map, inject it using the original key name
        elif self.cfg.file_name_column not in self.cfg.param_map:
            payload[self.cfg.file_name_column] = name_to_send
        
        if not self.cfg.api_filename_param and self.cfg.file_name_column not in self.cfg.param_map:
            self.logger.warning(
                f"api_filename_param is not set and the filename column '{self.cfg.file_name_column}' is not in param_map. "
                f"The key table's filename column name is being used as the payload key, which may be incorrect."
            )

        return payload

    def _execute_single_request(self, key_name: str, payload: Dict[str, Any]) -> Optional[requests.Response]:
        """Execute a single HTTP request with retries.

        Args:
            key_name (str): The filename key associated with this request.
            payload (Dict[str, Any]): JSON payload or query parameters.

        Returns:
            Optional[requests.Response]: A successful response object, or None on failure after retries.
        """
        method = self.cfg.http_method.upper()
        with requests.Session() as session:
            for attempt in range(self.cfg.max_retries):
                try:
                    if method == "POST":
                        resp = session.post(self.cfg.api_endpoint, json=payload, headers=self.cfg.headers, timeout=self.cfg.timeout_seconds)
                    elif method == "GET":
                        resp = session.get(self.cfg.api_endpoint, params=payload, headers=self.cfg.headers, timeout=self.cfg.timeout_seconds)
                    else:
                        raise ValueError(f"Unsupported HTTP method: {method}")
                    resp.raise_for_status()
                    self.logger.debug("[api] 200 OK for key '%s'", key_name)
                    return resp
                except requests.RequestException as e:
                    self.logger.warning(
                        "API request failed for key '%s' (attempt %d/%d): %s",
                        key_name,
                        attempt + 1,
                        self.cfg.max_retries,
                        e,
                    )
                    # On final attempt, record details for orchestrator-level logging
                    if attempt >= self.cfg.max_retries - 1:
                        status_code: Optional[int] = None
                        corr_id: Optional[str] = None
                        try:
                            resp_obj = getattr(e, "response", None)
                            if resp_obj is not None:
                                status_code = getattr(resp_obj, "status_code", None)
                                try:
                                    corr_id = resp_obj.headers.get("X-Correlation-ID")  # type: ignore[assignment]
                                except Exception:
                                    corr_id = None
                        except Exception:
                            status_code = None
                            corr_id = None
                        self._record_last_failure(key_name, status_code, str(e), corr_id)
                    if attempt < self.cfg.max_retries - 1:
                        time.sleep(self.cfg.backoff_seconds)
        return None

    def _save_and_get_path(self, key_name: str, response: requests.Response) -> Optional[Tuple[Path, str]]:
        """Persist a fetched image and return its path and filename.

        Handles base64, raw-binary, or URL-indirect responses according to
        configuration. Ensures file name safety, extension normalization to TIF,
        and uniqueness within the target directory.

        Args:
            key_name (str): The key for which the image was requested.
            response (requests.Response): HTTP response returned by the API.

        Returns:
            Optional[Tuple[pathlib.Path, str]]: Resolved file path and saved filename on success; None on failure.
        """
        try:
            # Choose save directory: persistent vs transient
            if self.cfg.persist_downloads:
                save_dir = self.cfg.api_output_save_path
            else:
                if self._transient_dir is None:
                    # Initialize a transient directory once per fetcher
                    if self.cfg.transient_download_root:
                        base = Path(self.cfg.transient_download_root)
                        base.mkdir(parents=True, exist_ok=True)
                        # Use time-based unique folder
                        tname = f"api_transient_{int(time.time()*1000)}"
                        self._transient_dir = (base / tname)
                        self._transient_dir.mkdir(parents=True, exist_ok=True)
                    else:
                        self._transient_dir = Path(tempfile.mkdtemp(prefix="api_transient_"))
                save_dir = self._transient_dir

            save_dir.mkdir(parents=True, exist_ok=True)
            
            content_type = response.headers.get("Content-Type", "") or ""
            is_json_response = "json" in content_type.lower()
            
            json_data = None
            if is_json_response:
                try:
                    json_data = response.json()
                except json.JSONDecodeError:
                    self.logger.error("Failed to parse JSON response for key '%s'", key_name)
                    self._log_failure(key_name, getattr(response, "status_code", None), "Failed to parse JSON response")
                    return None
                # Detect explicit failure signals in JSON responses even when HTTP status is 2xx
                try:
                    if isinstance(json_data, dict):
                        err_msg = json_data.get("error")
                        success_val = json_data.get("success")
                        # Consider it a failure if an error is present AND no image payload is present
                        if err_msg and not any(
                            json_data.get(k) for k in [self.cfg.image_field, self.cfg.url_field, self.cfg.binary_field]
                        ):
                            self._log_failure(
                                key_name,
                                getattr(response, "status_code", None),
                                f"API error: {err_msg}",
                            )
                            return None
                        if isinstance(success_val, bool) and success_val is False:
                            self._log_failure(
                                key_name,
                                getattr(response, "status_code", None),
                                "API reported success=false",
                            )
                            return None
                except Exception:
                    # Never fail on diagnostics; proceed to field-based checks below
                    pass

            # Determine the output filename
            if self.cfg.file_name_field and json_data:
                output_filename = json_data.get(self.cfg.file_name_field, key_name)
            else:
                output_filename = key_name

            # Sanitize and ensure uniqueness
            base, ext = os.path.splitext(output_filename)
            # sanitize base to filesystem-friendly
            base = "".join(c for c in base if c.isalnum() or c in ('_', '-', '.', ' ')).strip() or "image"
            # Preserve extension when provided; infer when missing using Content-Type when possible
            ext_l = (ext or "").lower()
            if not ext_l:
                ct = (response.headers.get("Content-Type", "") or "").lower()
                if "tiff" in ct:
                    ext = ".tif"
                elif "jpeg" in ct or "jpg" in ct:
                    ext = ".jpg"
                elif "png" in ct:
                    ext = ".png"
                else:
                    ext = ".tif"  # safe default
            # Normalize to TIF for key-mode TIF-only contract when server returns non-TIF
            if (ext or "").lower() not in (".tif", ".tiff"):
                ext = ".tif"
            save_path = save_dir / f"{base}{ext}"
            counter = 1
            while save_path.exists():
                save_path = save_dir / f"{base}_{counter}{ext}"
                counter += 1
            output_filename = save_path.name

            if self.cfg.image_payload_type == "base64":
                if not json_data:
                    self._log_failure(key_name, getattr(response, "status_code", None), "Empty JSON response")
                    return None
                img_b64 = json_data.get(self.cfg.image_field)
                if not img_b64:
                    self.logger.warning("API response for key '%s' missing '%s'.", key_name, self.cfg.image_field)
                    self._log_failure(key_name, getattr(response, "status_code", None), f"Missing JSON field: {self.cfg.image_field}")
                    return None
                with open(save_path, "wb") as f:
                    f.write(base64.b64decode(img_b64))
            elif self.cfg.image_payload_type == "binary":
                if is_json_response:
                    if not json_data:
                        self._log_failure(key_name, getattr(response, "status_code", None), "Empty JSON response (binary mode)")
                        return None
                    # Assumes JSON carrier with a field containing binary data
                    img_bytes_data = json_data.get(self.cfg.binary_field)
                    if isinstance(img_bytes_data, str):
                        img_bytes = base64.b64decode(img_bytes_data)
                    else:
                        self.logger.warning(f"Binary data in JSON for key '{key_name}' is not a base64 string.")
                        self._log_failure(key_name, getattr(response, "status_code", None), f"Missing/invalid JSON field for binary: {self.cfg.binary_field}")
                        return None
                else: # Raw binary response
                    ct = response.headers.get("Content-Type", "")
                    if ct and not (ct.lower().startswith("image/") or "octet-stream" in ct.lower()):
                        self.logger.warning("API returned non-image content-type for key '%s': %s", key_name, ct)
                    img_bytes = response.content

                if not img_bytes:
                    self.logger.warning(f"No binary content for key '{key_name}'.")
                    self._log_failure(key_name, getattr(response, "status_code", None), "No binary content in response")
                    return None
                with open(save_path, "wb") as f:
                    f.write(img_bytes)
            elif self.cfg.image_payload_type == "url":
                if not json_data:
                    self._log_failure(key_name, getattr(response, "status_code", None), "Empty JSON response (url mode)")
                    return None
                img_url = json_data.get(self.cfg.url_field)
                if not img_url:
                    self.logger.warning(f"API response for key '{key_name}' missing '{self.cfg.url_field}'.")
                    self._log_failure(key_name, getattr(response, "status_code", None), f"Missing JSON field: {self.cfg.url_field}")
                    return None
                
                # Download the image from the URL with retry
                for attempt in range(self.cfg.max_retries):
                    try:
                        with requests.get(img_url, timeout=self.cfg.timeout_seconds, stream=True) as img_resp:
                            img_resp.raise_for_status()
                            
                            # Validate content type if possible
                            ct = img_resp.headers.get("Content-Type")
                            if ct and not ct.startswith("image/"):
                                self.logger.warning(f"URL for key '{key_name}' returned non-image content-type: {ct}")

                            with open(save_path, "wb") as f:
                                for chunk in img_resp.iter_content(chunk_size=8192):
                                    f.write(chunk)
                        break # Success
                    except requests.RequestException as e:
                        self.logger.warning("URL download failed for key '%s' (attempt %d/%d): %s", key_name, attempt + 1, self.cfg.max_retries, e)
                        if attempt < self.cfg.max_retries - 1:
                            time.sleep(self.cfg.backoff_seconds)
                else:
                    self.logger.error("Failed to download image from URL for key '%s' after multiple retries.", key_name)
                    self._log_failure(key_name, getattr(response, "status_code", None), "Failed to download image from URL after retries")
                    return None
            else:
                self.logger.error(f"Unsupported image_payload_type: {self.cfg.image_payload_type}")
                self._log_failure(key_name, getattr(response, "status_code", None), f"Unsupported image_payload_type: {self.cfg.image_payload_type}")
                return None

            self.logger.info("[api] saved '%s' for key '%s' -> %s", output_filename, key_name, save_path.resolve())
            return save_path.resolve(), output_filename
        except Exception as e:
            self.logger.error("Failed to save image for key '%s': %s", key_name, e)
            try:
                status = getattr(response, "status_code", None)  # type: ignore[name-defined]
            except Exception:
                status = None
            self._log_failure(key_name, status, f"Exception during save: {e}")
            return None

    def fetch_batch(self, filenames: List[str], rows_map: Dict[str, Dict[str, Any]]) -> Tuple[List[Path], Dict[str, str]]:
        """Fetch a batch of images from the API concurrently.

        Args:
            filenames (List[str]): List of filename keys to request.
            rows_map (Dict[str, Dict[str, Any]]): Mapping of key -> source row used to build payloads.

        Returns:
            Tuple[List[pathlib.Path], Dict[str, str]]: A tuple of (resolved image paths, saved_name->key mapping).
        """
        resolved_paths: List[Path] = []
        name_map_saved_to_key: Dict[str, str] = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.cfg.max_concurrency) as executor:
            future_to_key = {}
            total = 0
            success_count = 0
            error_count = 0
            for key_name in filenames:
                row = rows_map.get(key_name)
                if not row:
                    self.logger.warning(f"No row data found for key '{key_name}' in the key table.")
                    continue
                payload = self._prepare_request_payload(key_name, row)
                future = executor.submit(self._execute_single_request, key_name, payload)
                future_to_key[future] = key_name
                total += 1

            for future in concurrent.futures.as_completed(future_to_key):
                key_name = future_to_key[future]
                try:
                    response = future.result()
                    if response:
                        saved_info = self._save_and_get_path(key_name, response)
                        if saved_info:
                            path, saved_name = saved_info
                            resolved_paths.append(path)
                            name_map_saved_to_key[saved_name] = key_name
                            success_count += 1
                        else:
                            error_count += 1
                    else:
                        error_count += 1
                        # Log final HTTP failure using recorded details, if available
                        info = self._pop_last_failure(key_name)
                        if info:
                            self._log_failure(
                                key_name,
                                info.get("status_code"),  # type: ignore[arg-type]
                                str(info.get("reason")),
                                info.get("correlation_id"),
                            )
                        else:
                            self._log_failure(key_name, None, "HTTP request failed after retries")
                except Exception as exc:
                    error_count += 1
                    self.logger.error("API fetch generated an exception for key '%s': %s", key_name, exc)
        # Per-batch summary
        self.logger.info("API batch: %d/%d succeeded, %d errors", success_count, total, error_count)
        return resolved_paths, name_map_saved_to_key

@dataclass
class LocalFetchConfig:
    """Configuration for resolving filenames under local directories.

    Attributes:
        search_roots (List[Path]): Directories to search for files.
        recursive (bool): Whether to search recursively.
        allowed_extensions (Tuple[str, ...]): Allowed file extensions.
        resolve_without_extension (bool): Try adding extensions when missing.
        case_insensitive_match (bool): Case-insensitive filename matching.
        stop_on_first_match (bool): Stop searching after first match.
    """
    search_roots: List[Path]
    recursive: bool = True
    allowed_extensions: Tuple[str, ...] = (".tif", ".tiff")
    resolve_without_extension: bool = True
    case_insensitive_match: bool = True
    stop_on_first_match: bool = True

@dataclass
class NameMappingConfig:
    """Filename mapping strategy configuration.

    Attributes:
        enabled (bool): Enable filename tail mapping.
        debug_log (bool): Emit detailed mapping search logs.
        tail_len (int): Tail length used to map.
        insert_token (str): Token inserted before the tail.
        glob_suffix (str): Glob suffix pattern for mapped names.
        use_rglob_any_depth (bool): Use rglob for recursive matching.
        db_like_template (str): Template used for DB LIKE queries when enabled.
    """
    enabled: bool = True
    debug_log: bool = False
    tail_len: int = 5
    insert_token: str = "001"
    glob_suffix: str = "_*.tif"
    use_rglob_any_depth: bool = True
    db_like_template: str = "{prefix}{insert}{suffix}_%.tif"

class LocalFolderFetcher:
    """Resolve filenames into absolute Paths under the configured roots.

    - Attempts direct join and existence check
    - If resolve_without_extension: tries appending known extensions
    - If recursive: performs rglob; otherwise checks direct path only
    - Case-insensitive matching when enabled
    - Returns only existing files; logs missing ones
    """
    def __init__(self, cfg: LocalFetchConfig, logger: Optional[logging.Logger] = None, name_map: Optional[NameMappingConfig] = None) -> None:
        self.cfg = cfg
        self.logger = logger or _get_logger(__name__ + ".LocalFolderFetcher")
        self.name_map = name_map or NameMappingConfig()

    def _mapped_core(self, filename: str) -> Optional[str]:
        if not self.name_map.enabled:
            return None
        n = len(filename)
        t = max(0, int(self.name_map.tail_len))
        if n <= t:
            return None
        prefix = filename[: n - t]
        suffix = filename[n - t :]
        return f"{prefix}{self.name_map.insert_token}{suffix}"

    def _candidate_names(self, filename: str) -> List[str]:
        """Return candidate filenames to try for local resolution.

        Includes mapped variants (with allowed extensions) followed by original
        names. This improves robustness when on-disk names follow a tokenized
        suffix pattern (e.g., 001 insertion).

        Args:
            filename (str): Key filename from the key table.

        Returns:
            List[str]: Ordered candidate names for existence checks.
        """
        # Prefer mapped name(s) first, then original as fallback
        names: List[str] = []
        root, ext = os.path.splitext(filename)
        mapped = self._mapped_core(filename)
        if mapped:
            names.append(mapped)
            if (not ext) and self.cfg.resolve_without_extension:
                for e in self.cfg.allowed_extensions:
                    names.append(mapped + e)

        # Original as fallback
        names.append(filename)
        if (not ext) and self.cfg.resolve_without_extension:
            for e in self.cfg.allowed_extensions:
                names.append(filename + e)
        return names

    def _match_case(self, name: str, entries: List[str]) -> Optional[str]:
        """Return a case-corrected filename when case-insensitive matching is enabled.

        Args:
            name (str): Candidate filename to match.
            entries (List[str]): Directory entry names to match against.

        Returns:
            Optional[str]: Matched name with original casing, or None if not found.
        """
        if name in entries:
            return name
        if self.cfg.case_insensitive_match:
            lower_map = {e.lower(): e for e in entries}
            return lower_map.get(name.lower())
        return None

    def fetch_batch(self, filenames: List[str]) -> List[Path]:
        """Resolve a batch of filenames to absolute existing Paths.

        Applies direct checks, optional extension resolution, recursive globbing
        with mapping-aware patterns, and case-insensitive comparisons according
        to configuration.

        Args:
            filenames (List[str]): Filenames requested from the key table.

        Returns:
            List[pathlib.Path]: Paths that exist on disk and have TIF/TIFF extensions.
        """
        resolved: List[Path] = []
        for fname in filenames:
            found_path: Optional[Path] = None
            cand_names = self._candidate_names(fname)
            mapped_for_log = self._mapped_core(fname)

            if self.name_map.debug_log:
                self.logger.info("query image filename from table: %s", fname)
                self.logger.info("modified query image filename for search: %s", (mapped_for_log if mapped_for_log else fname))

            for root in self.cfg.search_roots:
                existing = list(root.iterdir()) if root.exists() and not self.cfg.recursive else []
                entries = [p.name for p in existing] if existing else []

                for cn in cand_names:
                    p = root / cn
                    if p.exists():
                        found_path = p.resolve()
                        break
                    m = self._match_case(cn, entries) if entries else None
                    if m:
                        found_path = (root / m).resolve()
                        break
                if found_path:
                    break

                if self.cfg.recursive and root.exists():
                    try:
                        # If name mapping is enabled and mapped core exists, try glob with suffix first
                        mapped = self._mapped_core(fname)
                        if mapped and self.name_map.glob_suffix:
                            pattern = mapped + self.name_map.glob_suffix
                            if self.name_map.debug_log:
                                self.logger.info("[local] trying pattern under '%s': %s", root, pattern)
                            iterator = root.rglob(pattern) if self.name_map.use_rglob_any_depth else root.glob(pattern)
                            for rp in iterator:
                                if rp.is_file():
                                    if self.name_map.debug_log:
                                        self.logger.info("[local] key='%s' mapped='%s' -> matched '%s'", fname, mapped, rp)
                                    found_path = rp.resolve()
                                    break

                        # Also try exact mapped + extension patterns (e.g., no underscore)
                        if mapped and not found_path:
                            for e in self.cfg.allowed_extensions:
                                pattern = mapped + e
                                if self.name_map.debug_log:
                                    self.logger.info("[local] trying pattern under '%s': %s", root, pattern)
                                iterator = root.rglob(pattern) if self.name_map.use_rglob_any_depth else root.glob(pattern)
                                for rp in iterator:
                                    if rp.is_file():
                                        if self.name_map.debug_log:
                                            self.logger.info("[local] key='%s' mapped='%s' -> matched '%s'", fname, mapped, rp)
                                        found_path = rp.resolve()
                                        break
                                if found_path:
                                    break

                        if not found_path:
                            if self.name_map.debug_log:
                                self.logger.info("[local] fallback exact names under '%s' for key='%s' candidates=%s", root, fname, cand_names[:4])
                            for cn in cand_names:
                                for rp in root.rglob("*"):
                                    if rp.is_file():
                                        if rp.name == cn or (self.cfg.case_insensitive_match and rp.name.lower() == cn.lower()):
                                            found_path = rp.resolve()
                                            break
                                if found_path:
                                    break
                    except Exception as e:
                        self.logger.debug("rglob error under %s: %s", root, e)

                if found_path and self.cfg.stop_on_first_match:
                    break

            if found_path is None:
                self.logger.warning("Missing file: %s", fname)
                if self.name_map.debug_log:
                    pattern_hint = (mapped_for_log + self.name_map.glob_suffix) if (mapped_for_log and self.name_map.glob_suffix) else None
                    self.logger.info(
                        "[local] not found details key='%s' mapped='%s' roots=%s pattern='%s'",
                        fname,
                        mapped_for_log,
                        [str(r) for r in self.cfg.search_roots],
                        pattern_hint,
                    )
            else:
                # Enforce TIF-only regardless of configured allowed_extensions
                if found_path.suffix.lower() in (".tif", ".tiff"):
                    resolved.append(found_path)
                else:
                    self.logger.warning(f"Skipping non-TIF file: {found_path}")
        return resolved

@dataclass
class DatabaseConfig:
    """Configuration for database-backed TIF retrieval.

    Attributes:
        driver (str): Database driver (informational).
        host (str): Database host.
        port (int): Database port.
        database (str): Database name.
        user (str): Username.
        password_env_var (str): Env var name for password.
        sslmode (str): SSL mode for connection.
        fetch_mode (str): 'path' or 'blob'.
        query_template (str): SQL with %(file_name)s placeholder.
        like_query_template (Optional[str]): Optional LIKE query fallback.
        path_column (str): Column containing file path in 'path' mode.
        blob_column (str): Column containing blob in 'blob' mode.
        blob_temp_dir (Optional[Path]): Temporary dir for blob files.
    """
    driver: str = "postgresql" # informational; psycopg2 used
    host: str = "127.0.0.1"
    port: int = 5432
    database: str = "postgres"
    user: str = "postgres"
    password_env_var: str = "POSTGRES_PASSWORD"
    sslmode: str = "prefer"
    # How to obtain the file for a given filename
    fetch_mode: str = "path" # "path" or "blob"
    query_template: str = "SELECT file_path FROM tif_documents WHERE file_name = %(file_name)s"
    like_query_template: Optional[str] = None
    path_column: str = "file_path"
    blob_column: str = "file_blob"
    blob_temp_dir: Optional[Path] = None

class DatabaseFetcher:
    """Fetch TIF files from a PostgreSQL/EDB database by filename.

    Two modes:
    - path: the query returns a path to a file on disk (path_column), which must exist
    - blob: the query returns a binary blob (blob_column), which is written to a temp dir

    The query_template must accept a %(file_name)s placeholder.
    """
    def __init__(self, cfg: DatabaseConfig, logger: Optional[logging.Logger] = None, name_map: Optional[NameMappingConfig] = None) -> None:
        self.cfg = cfg
        self.logger = logger or _get_logger(__name__ + ".DatabaseFetcher")
        self.name_map = name_map or NameMappingConfig()
        try:
            import psycopg2 # type: ignore
            self._psycopg2 = psycopg2
        except ImportError as e:
            raise RuntimeError("Database mode requires psycopg2. Please install it.") from e

    def _connect(self) -> 'psycopg2.extensions.connection':
        """Create a new database connection using psycopg2.

        Returns:
            psycopg2.extensions.connection: A live PostgreSQL connection from configured DSN.
        """
        password = os.getenv(self.cfg.password_env_var, "")
        conn = self._psycopg2.connect(
            host=self.cfg.host,
            port=self.cfg.port,
            dbname=self.cfg.database,
            user=self.cfg.user,
            password=password,
            sslmode=self.cfg.sslmode,
        )
        return conn

    def _mapped_core(self, filename: str) -> Optional[str]:
        if not self.name_map.enabled:
            return None
        n = len(filename)
        t = max(0, int(self.name_map.tail_len))
        if n <= t:
            return None
        prefix = filename[: n - t]
        suffix = filename[n - t :]
        return f"{prefix}{self.name_map.insert_token}{suffix}"

    def fetch_batch(self, filenames: List[str]) -> List[Tuple[str, Path]]:
        """Fetch TIF files for a batch of keys using database lookups.

        Depending on ``fetch_mode``, this resolves file system paths or writes
        blobs to temporary files and returns their paths.

        Args:
            filenames (List[str]): Filename keys to resolve.

        Returns:
            List[Tuple[str, Path]]: List of (requested_key, resolved_path) pairs.
        """
        resolved: List[Tuple[str, Path]] = []
        conn = self._connect()
        try:
            with conn.cursor() as cur:
                for fname in filenames:
                    try:
                        rows: List[tuple] = []
                        mapped = self._mapped_core(fname) if self.name_map.enabled else None

                        if self.name_map.debug_log:
                            self.logger.info("query image filename from table: %s", fname)
                            self.logger.info("modified query image filename for search: %s", (mapped if mapped else fname))

                        query = self.cfg.query_template
                        params = {"file_name": mapped or fname}

                        # Try exact match first
                        cur.execute(query, params)
                        rows = cur.fetchall()

                        # Fallback to LIKE query if configured and no exact match found
                        if not rows and self.cfg.like_query_template:
                            query = self.cfg.like_query_template
                            cur.execute(query, params)
                            rows = cur.fetchall()
                        
                        # Fallback to original filename if mapped was tried
                        if not rows and mapped:
                            params["file_name"] = fname
                            cur.execute(self.cfg.query_template, params)
                            rows = cur.fetchall()

                        if not rows:
                            self.logger.warning("No DB rows for filename (mapped-first): %s", fname)
                            continue

                        for row in rows:
                            if self.cfg.fetch_mode == "path":
                                try:
                                    if isinstance(row, dict):
                                        path_str = row.get(self.cfg.path_column)
                                    else:
                                        desc = [d[0] for d in cur.description]
                                        idx = desc.index(self.cfg.path_column) if self.cfg.path_column in desc else 0
                                        path_str = row[idx]
                                except Exception:
                                    path_str = row[0]

                                if not path_str:
                                    continue
                                p = Path(str(path_str)).resolve()
                                if p.exists():
                                    resolved.append((fname, p))
                                else:
                                    self.logger.warning("DB path not found on disk: %s", p)
                            else: # blob mode
                                try:
                                    desc = [d[0] for d in cur.description]
                                    idx = desc.index(self.cfg.blob_column) if self.cfg.blob_column in desc else 0
                                    blob = row[idx]
                                except Exception:
                                    blob = row[0]

                                if not blob:
                                    continue

                                temp_dir = self.cfg.blob_temp_dir or Path(tempfile.gettempdir()) / "db_tif_blobs"
                                temp_dir.mkdir(parents=True, exist_ok=True)
                                
                                # Sanitize filename for path
                                safe_fname = "".join(c for c in fname if c.isalnum() or c in ('_', '-')).rstrip()
                                out_path = (temp_dir / safe_fname).with_suffix(".tif")
                                
                                # Ensure uniqueness
                                counter = 1
                                while out_path.exists():
                                    out_path = (temp_dir / f"{safe_fname}_{counter}").with_suffix(".tif")
                                    counter += 1

                                with open(out_path, "wb") as f:
                                    f.write(blob)
                                resolved.append((fname, out_path.resolve()))
                    except Exception as e_q:
                        self.logger.error("DB query failed for %s: %s", fname, e_q)
        finally:
            try:
                conn.close()
            except Exception:
                pass
        return resolved

# ==============================================================================
# Orchestrator
# ==============================================================================
@dataclass
class OrchestratorSummary:
    status: str
    exit_code: int
    total_files_requested: int
    total_files_resolved: int
    total_batches: int
    batch_results: List[Dict[str, Any]]

class KeyDrivenOrchestrator:
    """Coordinates key-driven batching and invokes existing pipelines per batch."""
    def __init__(self, logger: Optional[logging.Logger] = None) -> None:
        self.logger = logger or _get_logger(__name__ + ".Orchestrator")

    def _load_main_config(self, path: Path) -> Dict[str, Any]:
        """Load the primary pipeline YAML configuration.

        Args:
            path (Path): Path to the main configuration file (image_similarity_config.yaml).

        Returns:
            Dict[str, Any]: Parsed configuration mapping.
        """
        if load_and_merge_configs is not None:
            return load_and_merge_configs(path)
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    def _load_key_config(self, path: Path) -> Dict[str, Any]:
        """Load the key-input YAML configuration for the orchestrator.

        Args:
            path (Path): Path to the key_input_config.yaml file.

        Returns:
            Dict[str, Any]: Parsed key-input configuration mapping.
        """
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    def _project_root(self) -> Path:
        """Compute the repository root relative to this file location.

        Returns:
            Path: Absolute path to the project root directory.
        """
        # external/key_input/key_input_orchestrator.py -> key_input/ -> external/ -> TifDoc_nd_image_similarity_search (project root)
        return Path(__file__).resolve().parents[2]

    def _key_input_cfg(self, kcfg: Dict[str, Any]) -> KeyInputConfig:
        """Translate raw key_input YAML to a typed KeyInputConfig.

        Args:
            kcfg (Dict[str, Any]): Full key_input configuration mapping from YAML.

        Returns:
            KeyInputConfig: Dataclass with validated and defaulted options.
        """
        ki = kcfg.get("key_input", {})
        return KeyInputConfig(
            input_table_path=Path(ki.get("input_table_path")),
            file_name_column=str(ki.get("file_name_column", "파일명")),
            format=str(ki.get("format", "auto")),
            json_array_field=ki.get("json_array_field"),
            json_records_is_lines=bool(ki.get("json_records_is_lines", False)),
            batch_size=int(ki.get("batch_size", 200)),
            deduplicate=bool(ki.get("deduplicate", True)),
            strip_whitespace=bool(ki.get("strip_whitespace", True)),
            key_in_xlsx_to_csv=bool(ki.get("key_in_xlsx_to_csv", True)),
            excel_sheet_name=ki.get("excel_sheet_name"),
            columns_for_results=ki.get("columns_for_results"),
        )

    def _prepare_local_fetcher(self, kcfg: Dict[str, Any]) -> LocalFolderFetcher:
        """Instantiate a LocalFolderFetcher from key_input configuration.

        Args:
            kcfg (Dict[str, Any]): Key-input configuration mapping.

        Returns:
            LocalFolderFetcher: Configured fetcher for on-disk TIF resolution.
        """
        local_cfg = kcfg.get("data_source", {}).get("local", {})
        roots = [Path(p) for p in local_cfg.get("search_roots", [])]
        lf_cfg = LocalFetchConfig(
            search_roots=roots,
            recursive=bool(local_cfg.get("recursive", True)),
            # TIF-only by default. If overridden, non-TIF matches will be skipped with a warning.
            allowed_extensions=tuple(local_cfg.get("allowed_extensions", [".tif", ".tiff"])),
            resolve_without_extension=bool(local_cfg.get("resolve_without_extension", True)),
            case_insensitive_match=bool(kcfg.get("key_input", {}).get("case_insensitive_match", True)),
            stop_on_first_match=bool(local_cfg.get("stop_on_first_match", True)),
        )
        nm_cfg = NameMappingConfig(**(kcfg.get("name_mapping") or {}))
        return LocalFolderFetcher(lf_cfg, self.logger, name_map=nm_cfg)

    def _prepare_db_fetcher(self, kcfg: Dict[str, Any]) -> DatabaseFetcher:
        """Instantiate a DatabaseFetcher from key_input configuration.

        Args:
            kcfg (Dict[str, Any]): Key-input configuration mapping.

        Returns:
            DatabaseFetcher: Configured fetcher for DB-backed TIF retrieval.
        """
        db_cfg = kcfg.get("data_source", {}).get("database", {})
        df_cfg = DatabaseConfig(
            driver=str(db_cfg.get("driver", "postgresql")),
            host=str(db_cfg.get("host", "127.0.0.1")),
            port=int(db_cfg.get("port", 5432)),
            database=str(db_cfg.get("database", "postgres")),
            user=str(db_cfg.get("user", "postgres")),
            password_env_var=str(db_cfg.get("password_env_var", "POSTGRES_PASSWORD")),
            sslmode=str(db_cfg.get("sslmode", "prefer")),
            fetch_mode=str(db_cfg.get("fetch_mode", "path")).lower(),
            query_template=str(db_cfg.get("query_template", "SELECT file_path FROM tif_documents WHERE file_name = %(file_name)s")),
            like_query_template=db_cfg.get("like_query_template"),
            path_column=str(db_cfg.get("path_column", "file_path")),
            blob_column=str(db_cfg.get("blob_column", "file_blob")),
            blob_temp_dir=Path(db_cfg.get("blob_temp_dir")) if db_cfg.get("blob_temp_dir") else None,
        )
        nm_cfg = NameMappingConfig(**(kcfg.get("name_mapping") or {}))
        return DatabaseFetcher(df_cfg, self.logger, name_map=nm_cfg)

    def _prepare_api_fetcher(self, kcfg: Dict[str, Any]) -> 'ApiFetcher':
        """Instantiate an ApiFetcher from key_input configuration.

        Validates presence of a filename parameter mapping and response mapping
        requirements for the selected payload type.

        Args:
            kcfg (Dict[str, Any]): Key-input configuration mapping.

        Returns:
            ApiFetcher: Configured API fetcher.

        Raises:
            RuntimeError: When mandatory API configuration is missing or invalid.
        """
        api_cfg = kcfg.get("data_source", {}).get("api", {})
        req = api_cfg.get("request_mapping", {}) or {}
        param_map = req.get("param_map", {}) or {}
        param_cast = req.get("param_cast", {}) or {}
        resp = api_cfg.get("response_mapping", {}) or {}

        # Determine required columns and declared types for default casting hints
        sqlite_declared_types: Dict[str, str] = {}
        disk_enabled = False
        try:
            if derive_required_columns and derive_declared_types:
                required_cols = derive_required_columns(kcfg)
                sqlite_declared_types = derive_declared_types(kcfg, required_cols)
            disk_enabled = bool(kcfg.get("key_input", {}).get("disk_backed_index", {}).get("enabled", False))
        except Exception:
            sqlite_declared_types = {}
            disk_enabled = False

        cfg = ApiConfig(
            api_endpoint=str(api_cfg.get("api_endpoint")),
            http_method=str(api_cfg.get("http_method", "POST")),
            headers=api_cfg.get("headers") or {},
            timeout_seconds=int(api_cfg.get("timeout_seconds", 30)),
            max_retries=int(api_cfg.get("retry", {}).get("max_retries", 3)),
            backoff_seconds=int(api_cfg.get("retry", {}).get("backoff_seconds", 2)),
            max_concurrency=int(api_cfg.get("max_concurrency", 4)),
            param_map=param_map,
            api_filename_param=req.get("api_filename_param"),
            send_mapped_filename=bool(req.get("send_mapped_filename", False)),
            param_cast=param_cast,
            sqlite_declared_types=sqlite_declared_types,
            disk_index_enabled=disk_enabled,
            persist_downloads=bool(api_cfg.get("persist_downloads", False)),
            transient_download_root=Path(api_cfg.get("transient_download_root")).resolve() if api_cfg.get("transient_download_root") else None,
            file_name_column=str(kcfg.get("key_input", {}).get("file_name_column", "파일명")),
            image_payload_type=str(resp.get("image_payload_type", "base64")),
            image_field=str(resp.get("image_field", "image_b64")),
            url_field=str(resp.get("url_field", "image_url")),
            binary_field=str(resp.get("binary_field", "image_bytes")),
            file_name_field=resp.get("file_name_field"),
            api_output_save_path=Path(api_cfg.get("api_output_save_path", "./api_fetched_images")),
        )
        nm_cfg = NameMappingConfig(**(kcfg.get("name_mapping") or {}))

        if not cfg.api_endpoint or str(cfg.api_endpoint).strip().lower() in ("", "none"):
            raise RuntimeError("API config must set data_source.api.api_endpoint.")

        # Validate API config: require explicit filename parameter mapping to avoid ambiguous payloads
        if not cfg.api_filename_param and cfg.file_name_column not in cfg.param_map:
            raise RuntimeError(
                "API config is missing a mapping for the filename parameter. "
                "Set data_source.api.request_mapping.api_filename_param or include the key_input.file_name_column in request_mapping.param_map."
            )

        # Validate response mapping fields for selected payload type
        if cfg.image_payload_type == "base64" and not cfg.image_field:
            raise RuntimeError("API response_mapping.image_field must be set when image_payload_type='base64'.")
        if cfg.image_payload_type == "url" and not cfg.url_field:
            raise RuntimeError("API response_mapping.url_field must be set when image_payload_type='url'.")

        return ApiFetcher(cfg, self.logger, name_map=nm_cfg)

    def _append_key_table_log(
        self,
        input_path: Path,
        total_requested: int,
        total_resolved: int,
        run_started_at: float,
    ) -> None:
        """Append a lightweight summary line to the key table log.

        The intent is to avoid any per-row logging that could slow the pipeline.

        Args:
            input_path (Path): The key table file path that was processed.
            total_requested (int): Count of keys requested from the key table.
            total_resolved (int): Count of successfully resolved/downloaded files.
            run_started_at (float): POSIX timestamp when the run started.
        """
        try:
            env_log_dir = os.getenv("APP_LOG_DIR", "/home/appuser/app/logs")
            log_dir: Path = (Path(env_log_dir) / "key_input").resolve()
            log_dir.mkdir(parents=True, exist_ok=True)
            log_file: Path = log_dir / "key_table_logs.log"

            end_ts = time.time()
            duration_s = max(0.0, end_ts - float(run_started_at))
            failed = max(0, int(total_requested) - int(total_resolved))
            # Local time ISO for readability; timezone-aware if available
            ts = datetime.now().astimezone().isoformat(timespec="seconds")

            line = (
                f"{ts} | file={str(input_path)} | rows={int(total_requested)} | "
                f"success={int(total_resolved)} | failed={failed} | duration_s={duration_s:.3f}\n"
            )
            with log_file.open("a", encoding="utf-8") as f:
                f.write(line)
        except Exception:
            # Logging should never break the pipeline
            pass

    def _postprocess_input_table(
        self,
        kcfg: Dict[str, Any],
        input_table_path: Path,
        total_requested: int,
        total_resolved: int,
    ) -> None:
        """Backup or delete the processed key map file, based on configuration.

        The toggle is read from key_input.backup_processed_input_table.
        When True, the file is renamed by appending ".bkup" to its current name
        When True, the file is renamed by appending ".bkup" to its current name
        (e.g., filtered_rows.xlsx.bkup). When False, the file is deleted.

        This is only attempted if at least one row was processed.

        Args:
            kcfg (Dict[str, Any]): The loaded key_input YAML configuration dict.
            input_table_path (Path): Absolute or relative path to the input table.
            total_requested (int): Count of keys requested from the key table.
            total_resolved (int): Count of successfully resolved/downloaded files.
        """
        try:
            if int(total_requested) <= 0:
                return
            if not input_table_path:
                return
            p = Path(input_table_path)
            if not p.exists():
                return

            ki = kcfg.get("key_input", {}) if isinstance(kcfg, dict) else {}
            # Tri-state policy:
            #   - True  => backup (append ".bkup")
            #   - False => delete
            #   - Any other value (including null/None) => leave file unchanged (useful for debugging)
            raw_policy = ki.get("backup_processed_input_table", True)
            if isinstance(raw_policy, bool):
                if raw_policy:
                    # Append .bkup after extension
                    backup_path = Path(str(p) + ".bkup")
                    try:
                        # If a previous backup exists, remove it to avoid errors on rename
                        if backup_path.exists():
                            backup_path.unlink()
                    except Exception:
                        pass
                    p.rename(backup_path)
                    self.logger.info("Backed up processed key table -> %s", backup_path)
                else:
                    try:
                        p.unlink()
                        self.logger.info("Deleted processed key table -> %s", p)
                    except Exception as e_del:
                        self.logger.warning("Failed to delete processed key table '%s': %s", p, e_del)
            else:
                self.logger.info(
                    "backup_processed_input_table is non-boolean (%r); leaving key table unchanged.",
                    raw_policy,
                )
        except Exception:
            # Never break the pipeline on best-effort housekeeping
            pass

    def run(self, main_config_path: Union[str, Path]) -> Optional[Dict[str, Any]]:
        """Run key-driven pipeline if enabled. Return None if not applicable."""
        main_config_path = Path(main_config_path)
        main_config = self._load_main_config(main_config_path)

        input_mode = (main_config.get("input_mode") or {})
        # Default to 'key'. If an unsupported value is provided, fallback to 'key' with a warning.
        _raw_doc_mode = input_mode.get("doc_input_start", "key")
        start_mode = str(_raw_doc_mode).strip().lower()
        if start_mode not in {"tif", "key"}:
            self.logger.warning(
                "input_mode.doc_input_start='%s' is not supported; falling back to default 'key' input mode.",
                start_mode,
            )
            start_mode = "key"

        if start_mode != "key":
            self.logger.info("doc_input_start != 'key' (value: %s). Skipping key-driven pipeline.", start_mode)
            return None

        # Load key config
        key_cfg_path = Path(input_mode.get(
            "key_input_config_path", "external/key_input/key_input_config.yaml"
        ))
        if not key_cfg_path.is_absolute():
            key_cfg_path = (self._project_root() / key_cfg_path).resolve()
        kcfg = self._load_key_config(key_cfg_path)

        # Per-module log maintenance for key_input (configurable)
        try:
            lcfg = (kcfg.get("logging") or {}) if isinstance(kcfg, dict) else {}
            backup_logs = bool(lcfg.get("backup_logs", False))
            remove_logs_days = int(lcfg.get("remove_logs_days", 7))
            logs_dir = (self._project_root() / "external" / "key_input" / "logs").resolve()
            # Only consider key_table_logs* files in this module's logs
            maintain_log_files(logs_dir, stem="key_table_logs", remove_logs_days=remove_logs_days, backup_logs=backup_logs)
        except Exception:
            # Housekeeping must never impact the pipeline
            pass

        # Record run start time for lightweight log metrics
        run_start_time: float = time.time()

        # Build loader and determine source mode
        ki_cfg = self._key_input_cfg(kcfg)
        # If the key table is absent (e.g., previously processed and backed up/deleted),
        # short-circuit with a no-op so schedulers can run safely with no new work.
        if not Path(ki_cfg.input_table_path).exists():
            self.logger.info("Key table not found at '%s'; skipping key-driven pipeline run.", ki_cfg.input_table_path)
            return None
        loader = KeyInputLoader(ki_cfg, self.logger)
        mode = str(kcfg.get("data_source", {}).get("mode", "api")).lower()

        total_requested = 0
        total_resolved = 0
        batch_results: List[Dict[str, Any]] = []

        it: Iterator[str] = loader.iter_filenames()
        if ki_cfg.deduplicate:
            seen: set[str] = set()
            original_it = it # preserve original generator to avoid self-referential iteration
            def dedup_gen() -> Iterator[str]:
                for name in original_it:
                    if name in seen:
                        continue
                    seen.add(name)
                    yield name
            it = dedup_gen()

        # Determine required columns for disk-backed index
        required_columns: List[str] = []
        declared_types: Dict[str, str] = {}
        disk_cfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}) or {})
        disk_enabled = bool(disk_cfg.get("enabled", False)) and str(disk_cfg.get("backend", "sqlite")).lower() == "sqlite"
        if derive_required_columns and derive_declared_types:
            try:
                required_columns = derive_required_columns(kcfg)
                declared_types = derive_declared_types(kcfg, required_columns)
            except Exception as e:
                self.logger.debug("derive_* failed: %s", e)

        # If disk-backed index is enabled, ensure it exists and is up-to-date
        if disk_enabled and is_index_up_to_date and build_sqlite_index:
            # Honor global override for base directory when explicit db_path is not set in key config
            try:
                sqlite_cfg_current = (kcfg.get("key_input", {}).get("disk_backed_index", {}).get("sqlite") or {})
                has_explicit_db_path = bool(sqlite_cfg_current.get("db_path"))
                if not has_explicit_db_path:
                    qei = (main_config.get("search_task", {}).get("query_embed_index", {}) or {})
                    base_override = qei.get("output_path_query_embed_index")
                    if base_override:
                        # Deep copy kcfg and set db_path for this run only (does not persist to disk)
                        kcfg = json.loads(json.dumps(kcfg))
                        kcfg.setdefault("key_input", {}).setdefault("disk_backed_index", {}).setdefault("sqlite", {})["db_path"] = str(Path(base_override) / "key_table_index.sqlite")
            except Exception:
                pass
            # Support ephemeral (non-persistent) DBs by overriding db_path to a temp location
            persist_index = bool(disk_cfg.get("persist_disk_backed_index", True))
            if not persist_index:
                # Create a temp DB path per run
                temp_dir = Path(tempfile.mkdtemp(prefix="key_index_"))
                temp_db = temp_dir / "key_table_index.sqlite"
                # Mutate kcfg at runtime only for this run; this does not touch the YAML on disk
                kcfg = json.loads(json.dumps(kcfg))  # deep copy
                ddcfg = kcfg.setdefault("key_input", {}).setdefault("disk_backed_index", {})
                ddcfg.setdefault("sqlite", {})["db_path"] = str(temp_db)

            force_rebuild = bool(disk_cfg.get("rebuild_policy", {}).get("force_rebuild", False))
            up_to_date = False
            reason = "force_rebuild"
            if not force_rebuild:
                try:
                    up_to_date, reason = is_index_up_to_date(kcfg, required_columns)
                except Exception as e_up:
                    up_to_date, reason = False, f"check_failed:{e_up}"
            if not up_to_date:
                self.logger.info("[disk-index] rebuilding SQLite index (reason=%s)", reason)
                try:
                    build_sqlite_index(kcfg, required_columns)
                except Exception as e_build:
                    raise RuntimeError(f"Failed to build SQLite key index: {e_build}") from e_build
            else:
                self.logger.info("[disk-index] index up-to-date; reuse existing DB")
            # Log important runtime settings
            sqlite_cfg = (kcfg.get("key_input", {}).get("disk_backed_index", {}).get("sqlite") or {})
            db_path = sqlite_cfg.get("db_path", "instance/key_table_index.sqlite")
            self.logger.info(
                "Disk-backed key map enabled (backend=sqlite) db=%s case_insensitive_keys=%s index_chunk_size=%s persist=%s",
                db_path,
                bool(disk_cfg.get("case_insensitive_keys", True)),
                int(disk_cfg.get("index_chunk_size", 10000)),
                persist_index,
            )

            # If ephemeral, schedule cleanup
            if not persist_index:
                def _cleanup_temp_db(path: Path) -> None:
                    try:
                        if path.exists():
                            try:
                                path.unlink()
                            except Exception:
                                pass
                            # try removing parent
                            try:
                                path.parent.rmdir()
                            except Exception:
                                pass
                    except Exception:
                        pass
                # Ensure removal at process end by attaching to object lifetime
                self._temp_db_path = Path(db_path)
                self._temp_db_cleanup = _cleanup_temp_db

        # Load in-memory rows map only when required and disk-backed index is disabled
        rows_map: Optional[Dict[str, Dict[str, Any]]] = None
        needs_rows_map = ((mode == "api") or bool(ki_cfg.columns_for_results)) and not disk_enabled
        if needs_rows_map:
            try:
                rows_map = loader.load_rows_map()
            except Exception as e_rows:
                rows_map = None
                self.logger.warning("Key table row map unavailable: %s", e_rows)

        for batch in _ensure_iterable_chunks(it, max(1, ki_cfg.batch_size)):
            total_requested += len(batch)
            self.logger.info("Processing filename batch of size %d", len(batch))
            per_query_meta_map: Dict[str, Dict[str, Any]] = {}
            # Map of saved/matched filename (as processed in this batch) -> original key filename (normalized to .tif if missing)
            query_name_map: Dict[str, str] = {}

            if mode == "local":
                lfetch = self._prepare_local_fetcher(kcfg)
                resolved_paths = []
                # For enrichment, prefetch rows for the batch using disk index when enabled
                lookup_map: Dict[str, Dict[str, Any]] = {}
                if ki_cfg.columns_for_results and disk_enabled and lookup_rows_by_keys:
                    t0 = time.time()
                    try:
                        lookup_map = lookup_rows_by_keys(kcfg, batch, required_columns)
                    finally:
                        self.logger.info("[disk-index] batch lookup: %d keys in %.3fs", len(batch), time.time() - t0)

                for fname in batch:
                    paths = lfetch.fetch_batch([fname])
                    if paths:
                        p = paths[0]
                        resolved_paths.append(p)
                        # Build saved->original name mapping (normalize to .tif if key lacked extension)
                        orig_with_tif = fname if Path(fname).suffix else f"{fname}.tif"
                        # Only map TIF query documents; plain images are handled separately
                        if p.suffix.lower() in (".tif", ".tiff"):
                            query_name_map[p.name] = orig_with_tif
                        if ki_cfg.columns_for_results:
                            if disk_enabled:
                                row = lookup_map.get(fname)
                            else:
                                row = rows_map.get(fname) if rows_map else None
                            if row:
                                per_query_meta_map[orig_with_tif] = {c: row.get(c) for c in ki_cfg.columns_for_results}
            elif mode in ("database", "db"):
                self.logger.warning("data_source.mode 'database' is deprecated. Please switch to 'api'. Proceeding with DB for compatibility.")
                dbfetch = self._prepare_db_fetcher(kcfg)
                resolved_pairs = dbfetch.fetch_batch(batch)
                resolved_paths = [p for _, p in resolved_pairs]
                # Build saved->original mapping and metadata keyed by original (.tif if missing)
                for key, p in resolved_pairs:
                    orig_with_tif = key if Path(key).suffix else f"{key}.tif"
                    if p.suffix.lower() in (".tif", ".tiff"):
                        query_name_map[p.name] = orig_with_tif
                if rows_map and ki_cfg.columns_for_results:
                    for key, p in resolved_pairs:
                        row = rows_map.get(key)
                        if row:
                            orig_with_tif = key if Path(key).suffix else f"{key}.tif"
                            per_query_meta_map[orig_with_tif] = {c: row.get(c) for c in ki_cfg.columns_for_results}
            elif mode == "api":
                apifetch = self._prepare_api_fetcher(kcfg)
                # Configure failed-key-request logging from main config (default: enabled)
                try:
                    enable_flag = bool(main_config.get("logging", {}).get("enable_failed_key_request_logging", True))
                    setattr(apifetch, "enable_failed_key_request_logging", enable_flag)
                except Exception:
                    pass
                # Build per-batch rows map via disk index when enabled; else use in-memory rows_map
                if disk_enabled and lookup_rows_by_keys:
                    t0 = time.time()
                    batch_rows = lookup_rows_by_keys(kcfg, batch, required_columns)
                    self.logger.info("[disk-index] batch lookup: found %d/%d rows in %.3fs", len(batch_rows), len(batch), time.time() - t0)
                    if len(batch_rows) < len(batch):
                        missing = [k for k in batch if k not in batch_rows]
                        for m in missing:
                            self.logger.warning("No row data found for key '%s' in the key index.", m)
                    rows_map_effective = batch_rows
                else:
                    if not rows_map:
                        raise RuntimeError("API mode requires loading the key table rows to build payloads.")
                    rows_map_effective = rows_map

                resolved_paths, name_map_saved_to_key = apifetch.fetch_batch(batch, rows_map_effective)
                self.logger.info("API resolved %d/%d in this batch", len(resolved_paths), len(batch))
                # Build saved->original mapping (.tif if the key lacked extension)
                for saved_name, key_name in (name_map_saved_to_key or {}).items():
                    orig_with_tif = key_name if Path(key_name).suffix else f"{key_name}.tif"
                    if Path(saved_name).suffix.lower() in (".tif", ".tiff"):
                        query_name_map[saved_name] = orig_with_tif
                if ki_cfg.columns_for_results:
                    for saved_name, key_name in name_map_saved_to_key.items():
                        if disk_enabled:
                            row = rows_map_effective.get(key_name)
                        else:
                            row = rows_map.get(key_name) if rows_map else None
                        if row:
                            orig_with_tif = key_name if Path(key_name).suffix else f"{key_name}.tif"
                            per_query_meta_map[orig_with_tif] = {c: row.get(c) for c in ki_cfg.columns_for_results}
            else:
                raise ValueError(f"Unsupported data_source.mode: {mode}")

            # TIF-only: filter resolved items to .tif/.tiff
            tif_exts = {".tif", ".tiff"}
            tif_resolved_paths = [p for p in resolved_paths if p.suffix.lower() in tif_exts]

            total_resolved += len(tif_resolved_paths)

            # If no TIFs were resolved, skip this batch
            if not tif_resolved_paths:
                self.logger.warning("No .tif/.tiff files found in this batch. Skipping.")
                batch_results.append({
                    "mode": mode,
                    "requested": len(batch),
                    "resolved": 0,
                    "status": "skipped:no_files",
                })
                continue

            if execute_tif_batch_search_workflow is None:
                raise RuntimeError("core_engine workflow is not importable. Ensure project structure is intact.")

            tmpdir = Path(tempfile.mkdtemp(prefix="key_batch_"))
            try:
                # Copy each resolved TIF into a temp folder to reuse the existing batch folder pipeline semantics
                for p in tif_resolved_paths:
                    dst = tmpdir / p.name
                    try:
                        shutil.copy2(p, dst)
                    except Exception:
                        shutil.copy(p, dst)

                # Load main config fresh per batch (no extra overrides from key module)
                if load_and_merge_configs is not None:
                    cfg = load_and_merge_configs(main_config_path)
                else:
                    with main_config_path.open("r", encoding="utf-8") as f:
                        cfg = yaml.safe_load(f)

                # Inject key enrichment mapping for this batch
                if ki_cfg.columns_for_results:
                    cfg['key_enrichment'] = {
                        'columns_for_results': ki_cfg.columns_for_results,
                        'per_query_metadata_map': per_query_meta_map
                    }
                # Inject saved->original query name mapping for this batch (TIF-only)
                cfg['key_input_runtime'] = {
                    'query_name_map': query_name_map,
                }

                proj_root = self._project_root()

                # Optionally override indexing input to use the same temp folder when running in key mode
                key_mode_cfg = kcfg.get("key_mode", {}) or {}
                if bool(key_mode_cfg.get("use_tmp_for_indexing_input", False)):
                    try:
                        idx_cfg = cfg.setdefault('indexing_task', {})
                        # In key-mode TIF document workflows, override ONLY the TIF indexing input folder.
                        # Do not override image_folder_to_index (that is reserved for image-only indexing flows).
                        idx_cfg['input_tif_folder_for_indexing'] = str(tmpdir)
                    except Exception:
                        pass

                # Execute the existing TIF batch search workflow
                result = execute_tif_batch_search_workflow(
                    config=cfg,
                    project_root=proj_root,
                    input_folder_override=tmpdir,
                )
                batch_results.append({
                    "mode": mode,
                    "requested": len(batch),
                    "resolved": len(tif_resolved_paths),
                    "result": result,
                    "status": "ok" if int(result.get("exit_code", 1)) == 0 else "error"
                })
            finally:
                try:
                    shutil.rmtree(tmpdir)
                except Exception:
                    pass
                # Cleanup transient API downloads if configured
                if mode == "api":
                    try:
                        # apifetch is defined in the 'api' branch
                        if 'apifetch' in locals() and not apifetch.cfg.persist_downloads:
                            for p in resolved_paths:
                                try:
                                    p.unlink(missing_ok=True)  # type: ignore[arg-type]
                                except Exception:
                                    pass
                            # Attempt to remove now-empty parent folder(s)
                            try:
                                parents = {p.parent for p in resolved_paths}
                                for d in parents:
                                    try:
                                        d.rmdir()
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                    except Exception:
                        pass

        # Best-effort: append a single, lightweight summary log line and post-process the key table file
        try:
            self._append_key_table_log(
                input_path=ki_cfg.input_table_path,
                total_requested=total_requested,
                total_resolved=total_resolved,
                run_started_at=run_start_time,
            )
        finally:
            try:
                self._postprocess_input_table(
                    kcfg=kcfg,
                    input_table_path=ki_cfg.input_table_path,
                    total_requested=total_requested,
                    total_resolved=total_resolved,
                )
            except Exception:
                pass

        any_error = any(br.get("status") == "error" for br in batch_results)
        summary: OrchestratorSummary = OrchestratorSummary(
            status="success" if not any_error else "partial_error",
            exit_code=0 if not any_error else 1,
            total_files_requested=total_requested,
            total_files_resolved=total_resolved,
            total_batches=len(batch_results),
            batch_results=batch_results,
        )

        # Cleanup ephemeral disk-backed index if one was created
        try:
            temp_db_path = getattr(self, "_temp_db_path", None)
            temp_db_cleanup = getattr(self, "_temp_db_cleanup", None)
            if temp_db_path and temp_db_cleanup:
                temp_db_cleanup(temp_db_path)
        except Exception:
            pass

        return {
            "status": summary.status,
            "exit_code": summary.exit_code,
            "total_files_requested": summary.total_files_requested,
            "total_files_resolved": summary.total_files_resolved,
            "total_batches": summary.total_batches,
            "batch_results": summary.batch_results,
        }

# ==============================================================================
# Public API
# ==============================================================================
def run_key_input_pipeline(main_config_path: Union[str, Path]) -> Optional[Dict[str, Any]]:
    """Convenience wrapper for one-shot execution."""
    orch = KeyDrivenOrchestrator()
    return orch.run(main_config_path)


def run_key_input_pipeline_for_index(main_config_path: Union[str, Path]) -> Optional[Dict[str, Any]]:
    """Key-driven index builder: fetch TIFs per key-batch and build the index.

    This mirrors the key-driven search pipeline but, for each staged batch folder,
    calls build_index_from_tif_folder_workflow instead of executing the search.

    Args:
        main_config_path: Path to the primary YAML config (image_similarity_config.yaml).

    Returns:
        Optional[Dict[str, Any]]: Summary dict with per-batch results and counts, or None when
        input_mode.doc_input_start != 'key'.
    """
    logger = _get_logger(__name__ + ".index_builder")
    # Load configs
    if load_and_merge_configs is not None:
        main_config: Dict[str, Any] = load_and_merge_configs(main_config_path)  # type: ignore[arg-type]
    else:
        with Path(main_config_path).open("r", encoding="utf-8") as f:
            main_config = yaml.safe_load(f)

    input_mode: Dict[str, Any] = (main_config.get("input_mode") or {})
    start_mode = str(input_mode.get("doc_input_start", "key")).strip().lower()
    if start_mode not in {"tif", "key"}:
        logger.warning("Unsupported doc_input_start=%s; falling back to 'key'", start_mode)
        start_mode = "key"
    if start_mode != "key":
        return None

    # Determine key_input config path
    key_cfg_path = Path(input_mode.get("key_input_config_path", "external/key_input/key_input_config.yaml"))
    if not key_cfg_path.is_absolute():
        key_cfg_path = (Path(__file__).resolve().parents[2] / key_cfg_path).resolve()
    if not key_cfg_path.exists():
        raise FileNotFoundError(f"Key input config not found: {key_cfg_path}")

    with key_cfg_path.open("r", encoding="utf-8") as f:
        key_cfg: Dict[str, Any] = yaml.safe_load(f)

    ki_cfg = KeyInputConfig(
        input_table_path=Path((key_cfg.get("key_input") or {}).get("input_table_path")),
        file_name_column=str((key_cfg.get("key_input") or {}).get("file_name_column", "파일명")),
        format=str((key_cfg.get("key_input") or {}).get("format", "auto")),
        json_array_field=(key_cfg.get("key_input") or {}).get("json_array_field"),
        json_records_is_lines=bool((key_cfg.get("key_input") or {}).get("json_records_is_lines", False)),
        batch_size=int((key_cfg.get("key_input") or {}).get("batch_size", 200)),
        deduplicate=bool((key_cfg.get("key_input") or {}).get("deduplicate", True)),
        strip_whitespace=bool((key_cfg.get("key_input") or {}).get("strip_whitespace", True)),
        key_in_xlsx_to_csv=bool((key_cfg.get("key_input") or {}).get("key_in_xlsx_to_csv", True)),
        excel_sheet_name=(key_cfg.get("key_input") or {}).get("excel_sheet_name"),
        columns_for_results=(key_cfg.get("key_input") or {}).get("columns_for_results"),
    )

    data_mode = str((key_cfg.get("data_source") or {}).get("mode", "local")).lower()  # local|api|database
    disk_enabled = bool(((key_cfg.get("key_input") or {}).get("disk_backed_index") or {}).get("enabled", False))

    loader = KeyInputLoader(ki_cfg, logger)
    filenames = list(loader.iter_filenames())
    total_requested = len(filenames)

    # Optional global row map for API casting and enrichment
    rows_map: Optional[Dict[str, Dict[str, Any]]] = None
    # Always load rows_map for API mode so request payloads can be constructed,
    # even when disk-backed index is enabled. This avoids missing-row warnings.
    needs_rows_map = (data_mode == "api") or bool(ki_cfg.columns_for_results)
    if needs_rows_map:
        try:
            rows_map = loader.load_rows_map()
            if data_mode == "api" and not rows_map:
                logger.warning(
                    "Key-input rows map is empty; API mode requires row context for request payload mapping."
                )
        except Exception:
            rows_map = None

    # Prepare fetchers per mode
    local_fetcher: Optional[LocalFolderFetcher] = None
    api_fetcher: Optional[ApiFetcher] = None
    db_fetcher: Optional[DatabaseFetcher] = None

    if data_mode == "local":
        local_fetcher = KeyDrivenOrchestrator()._prepare_local_fetcher(key_cfg)
    elif data_mode == "api":
        api_fetcher = KeyDrivenOrchestrator()._prepare_api_fetcher(key_cfg)
    elif data_mode == "database":
        db_fetcher = KeyDrivenOrchestrator()._prepare_db_fetcher(key_cfg)
    else:
        raise ValueError(f"Unsupported data_source.mode: {data_mode}")

    if build_index_from_tif_folder_workflow is None:
        raise RuntimeError("core_engine workflow is not importable. Ensure project structure is intact.")

    # Project root for workflows
    project_root: Path = Path(__file__).resolve().parents[2]

    total_resolved = 0
    batch_results: List[Dict[str, Any]] = []

    for batch in _ensure_iterable_chunks(filenames, ki_cfg.batch_size):
        # Resolve/fetch files for this batch
        resolved_paths: List[Path] = []
        if data_mode == "local" and local_fetcher is not None:
            resolved_paths = local_fetcher.fetch_batch(batch)
        elif data_mode == "api" and api_fetcher is not None:
            rp, _ = api_fetcher.fetch_batch(batch, rows_map or {})
            resolved_paths = rp
        elif data_mode == "database" and db_fetcher is not None:
            kv = db_fetcher.fetch_batch(batch)
            resolved_paths = [p for (_k, p) in kv]

        total_resolved += len(resolved_paths)
        if not resolved_paths:
            logger.warning("No TIFs resolved for this batch; skipping index build call.")
            batch_results.append({"mode": data_mode, "requested": len(batch), "resolved": 0, "result": {"status": "success", "exit_code": 0, "message": "No files in batch."}})
            continue

        # Stage into a temporary folder for the core indexing workflow
        with tempfile.TemporaryDirectory(prefix="key_build_batch_") as tmpdir:
            tmp_dir = Path(tmpdir)
            for src in resolved_paths:
                try:
                    dst = tmp_dir / src.name
                    if src.resolve() != dst.resolve():
                        shutil.copy2(src, dst)
                except Exception as e:
                    logger.warning("Failed to stage file '%s': %s", src, e)
            # Invoke the indexing workflow for this staged batch
            try:
                res = build_index_from_tif_folder_workflow(main_config, project_root, input_tif_folder_override=tmp_dir)
            except Exception as e:
                res = {"status": "error", "exit_code": 1, "message": str(e)}
            batch_results.append({"mode": data_mode, "requested": len(batch), "resolved": len(resolved_paths), "result": res})

    summary: Dict[str, Any] = {
        "status": "success",
        "exit_code": 0,
        "total_files_requested": total_requested,
        "total_files_resolved": total_resolved,
        "total_batches": len(batch_results),
        "batch_results": batch_results,
    }
    return summary
